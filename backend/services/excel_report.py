"""
Excel report generation
"""

import pandas as pd
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.helpers import format_order_for_workers
from utils.workers import normalize_worker_name


def create_excel_report(data: List[dict], period: str, config: dict, for_workers: bool = False) -> bytes:
    """Create Excel report with proper formatting and formulas
    
    Args:
        for_workers: If True, creates simplified version for workers with hidden columns
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист_1"
    
    # Colors
    HEADER_BLUE = "4574A0"
    ROW_LIGHT_BLUE = "C6E2FF"
    DIAGNOSTIC_RED = "FF0000"
    YELLOW_TOTAL = "FFFF00"
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Styles - Arial font
    header_font = Font(name='Arial', bold=True, size=9, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    data_font = Font(name='Arial', size=9)
    param_font = Font(name='Arial', size=9)
    worker_font = Font(name='Arial', bold=True, size=9)
    worker_fill = PatternFill("solid", fgColor=ROW_LIGHT_BLUE)
    diagnostic_header_fill = PatternFill("solid", fgColor=DIAGNOSTIC_RED)
    yellow_fill = PatternFill("solid", fgColor=YELLOW_TOTAL)
    
    alignment_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Column widths
    column_widths = {
        'A': 55,
        'B': 12, 'C': 13, 'D': 12, 'E': 13,
        'F': 13, 'G': 15, 'H': 15, 'I': 14,
        'J': 12, 'K': 12, 'L': 14, 'M': 12, 'N': 14
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Hide columns for workers version
    if for_workers:
        ws.column_dimensions['B'].hidden = True
        ws.column_dimensions['C'].hidden = True
        ws.column_dimensions['G'].hidden = True
        ws.column_dimensions['H'].hidden = True
        ws.column_dimensions['I'].hidden = True
    
    # Parameter rows (1-3)
    ws['A1'] = "Параметры:"
    ws['A1'].font = param_font
    ws['A2'] = f"Период: {period}"
    ws['A2'].font = param_font
    ws['A3'] = f"Процент оплаты диагностики: {config['diagnostic_percent']}%"
    ws['A3'].font = param_font
    
    # Column headers (row 5)
    headers = [
        ("A", "Монтажник"),
        ("B", "Выручка итого"), ("C", "Выручка от услуг"),
        ("D", "Диагностика"), ("E", "Оплата диагностики"),
        ("F", "Выручка (выезд) специалиста"),
        ("G", "Доп. расходы (Оплата услуг помощников)"),
        ("H", "Сумма оплаты от услуг"),
        ("I", "Процент от выручки по услугам"),
        ("J", "Оплата бензина"), ("K", "Транспортные"),
        ("L", "Яндекс заправки"),  # NEW column
        ("M", "Итого"), ("N", "Диагностика -50%")
    ]
    
    for col_letter, header_text in headers:
        cell = ws[f"{col_letter}5"]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill if col_letter != "N" else diagnostic_header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
    
    ws.row_dimensions[5].height = 45
    
    # Group data by worker
    workers_data = {}
    for record in data:
        worker = record.get("worker", "").replace(" (оплата клиентом)", "")
        if worker not in workers_data:
            workers_data[worker] = {"regular": [], "client_payment": []}
        
        if record.get("is_client_payment"):
            workers_data[worker]["client_payment"].append(record)
        else:
            workers_data[worker]["regular"].append(record)
    
    current_row = 6
    
    def to_int(val):
        """Convert value to integer, return empty string if invalid"""
        if val is None or val == "" or (isinstance(val, float) and pd.isna(val)):
            return ""
        try:
            return int(round(float(val)))
        except:
            return ""
    
    for worker in sorted(workers_data.keys()):
        if not worker:
            continue
            
        worker_data = workers_data[worker]
        regular_rows = worker_data["regular"]
        client_rows = worker_data["client_payment"]
        
        # Worker name row
        worker_name_row = current_row
        cell = ws.cell(row=current_row, column=1, value=worker)
        cell.font = worker_font
        cell.fill = worker_fill
        cell.alignment = alignment_wrap
        cell.border = thin_border
        
        for col in range(2, 15):  # Extended to column 15 (N)
            c = ws.cell(row=current_row, column=col)
            c.fill = worker_fill
            c.border = thin_border
        
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        
        regular_start = current_row
        
        # Regular orders
        for record in regular_rows:
            if record.get("is_worker_total"):
                continue
            
            order_text = record.get("order", "")
            if for_workers:
                order_text = format_order_for_workers(order_text)
            
            cell = ws.cell(row=current_row, column=1, value=order_text)
            cell.font = data_font
            cell.alignment = alignment_wrap
            cell.border = thin_border
            
            for col, key in [(2, "revenue_total"), (3, "revenue_services"), (4, "diagnostic"),
                            (5, "diagnostic_payment"), (6, "specialist_fee"), (7, "additional_expenses"),
                            (8, "service_payment")]:
                val = to_int(record.get(key, ""))
                c = ws.cell(row=current_row, column=col, value=val if val != "" else None)
                c.font = data_font
                c.border = thin_border
            
            c = ws.cell(row=current_row, column=9, value=record.get("percent", ""))
            c.font = data_font
            c.border = thin_border
            
            fuel = to_int(record.get("fuel_payment", 0))
            transport = to_int(record.get("transport", 0))
            
            c = ws.cell(row=current_row, column=10, value=fuel if fuel else None)
            c.font = data_font
            c.border = thin_border
            
            c = ws.cell(row=current_row, column=11, value=transport if transport else None)
            c.font = data_font
            c.border = thin_border
            
            # Column 12 - Yandex Fuel (empty for individual orders)
            c = ws.cell(row=current_row, column=12)
            c.border = thin_border
            
            # Column 13 - Total (was 12)
            total_val = to_int(record.get("total", 0))
            c = ws.cell(row=current_row, column=13, value=total_val if total_val else None)
            c.font = data_font
            c.border = thin_border
            
            # Column 14 - Diagnostic -50% (was 13)
            c = ws.cell(row=current_row, column=14)
            c.border = thin_border
            
            current_row += 1
        
        regular_end = current_row - 1
        
        # Calculate sums for worker total row
        if regular_end >= regular_start:
            fuel_sum = sum(to_int(r.get("fuel_payment", 0)) or 0 for r in regular_rows if not r.get("is_worker_total"))
            transport_sum = sum(to_int(r.get("transport", 0)) or 0 for r in regular_rows if not r.get("is_worker_total"))
            
            c = ws.cell(row=worker_name_row, column=10, value=fuel_sum if fuel_sum else None)
            c.font = worker_font
            c = ws.cell(row=worker_name_row, column=11, value=transport_sum if transport_sum else None)
            c.font = worker_font
        
        # Client payment section
        client_name_row = None
        if client_rows:
            client_name_row = current_row
            cell = ws.cell(row=current_row, column=1, value=f"{worker} (оплата клиентом)")
            cell.font = worker_font
            cell.fill = worker_fill
            cell.alignment = alignment_wrap
            cell.border = thin_border
            
            for col in range(2, 15):  # Extended to column 15
                c = ws.cell(row=current_row, column=col)
                c.fill = worker_fill
                c.border = thin_border
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
            client_start = current_row
            
            for record in client_rows:
                if record.get("is_worker_total"):
                    continue
                
                order_text = record.get("order", "")
                if for_workers:
                    order_text = format_order_for_workers(order_text)
                
                cell = ws.cell(row=current_row, column=1, value=order_text)
                cell.font = data_font
                cell.alignment = alignment_wrap
                cell.border = thin_border
                
                for col, key in [(2, "revenue_total"), (3, "revenue_services"), (4, "diagnostic"),
                                (5, "diagnostic_payment"), (6, "specialist_fee"), (7, "additional_expenses"),
                                (8, "service_payment")]:
                    val = to_int(record.get(key, ""))
                    c = ws.cell(row=current_row, column=col, value=val if val != "" else None)
                    c.font = data_font
                    c.border = thin_border
                
                c = ws.cell(row=current_row, column=9, value=record.get("percent", ""))
                c.font = data_font
                c.border = thin_border
                
                for col in [10, 11, 12]:  # fuel, transport, yandex_fuel columns
                    ws.cell(row=current_row, column=col).border = thin_border
                
                # Column 13 - Total (was 12)
                total_val = to_int(record.get("total", 0))
                c = ws.cell(row=current_row, column=13, value=total_val if total_val else None)
                c.font = data_font
                c.border = thin_border
                
                # Column 14 - Diagnostic -50% (was 13)
                diag_50 = to_int(record.get("diagnostic_50", 0))
                c = ws.cell(row=current_row, column=14, value=diag_50 if diag_50 else None)
                c.font = data_font
                c.border = thin_border
                
                current_row += 1
            
            client_end = current_row - 1
            
            if client_end >= client_start:
                client_total_sum = sum(to_int(r.get("total", 0)) or 0 for r in client_rows if not r.get("is_worker_total"))
                client_diag50_sum = sum(to_int(r.get("diagnostic_50", 0)) or 0 for r in client_rows if not r.get("is_worker_total"))
                
                # Column 13 = Total, Column 14 = Diagnostic -50%
                c = ws.cell(row=client_name_row, column=13, value=client_total_sum if client_total_sum else None)
                c.font = worker_font
                c = ws.cell(row=client_name_row, column=14, value=client_diag50_sum if client_diag50_sum else None)
                c.font = worker_font
        
        # Main worker row Итого
        regular_total_sum = sum(to_int(r.get("total", 0)) or 0 for r in regular_rows if not r.get("is_worker_total"))
        client_diag50_sum_for_main = sum(to_int(r.get("diagnostic_50", 0)) or 0 for r in client_rows if not r.get("is_worker_total")) if client_rows else 0
        
        # Get Yandex Fuel deduction for this worker (from config)
        yandex_fuel_deduction = config.get("yandex_fuel", {}).get(worker, 0)
        
        if regular_end >= regular_start:
            if client_name_row:
                main_total = regular_total_sum - client_diag50_sum_for_main - yandex_fuel_deduction
            else:
                main_total = regular_total_sum - yandex_fuel_deduction
            
            # Column 12 - Yandex Fuel (only show if there's a deduction)
            if yandex_fuel_deduction:
                c = ws.cell(row=worker_name_row, column=12, value=-int(yandex_fuel_deduction))
                c.font = worker_font
            
            # Column 13 - Total
            c = ws.cell(row=worker_name_row, column=13, value=int(main_total) if main_total else None)
            c.font = worker_font
            c.fill = yellow_fill
            c.border = thin_border
        else:
            if client_name_row:
                main_total = -client_diag50_sum_for_main - yandex_fuel_deduction
                
                # Column 12 - Yandex Fuel
                if yandex_fuel_deduction:
                    c = ws.cell(row=worker_name_row, column=12, value=-int(yandex_fuel_deduction))
                    c.font = worker_font
                
                c = ws.cell(row=worker_name_row, column=13, value=int(main_total) if main_total else None)
            else:
                c = ws.cell(row=worker_name_row, column=13, value=0)
            c.font = worker_font
            c.fill = yellow_fill
            c.border = thin_border
        
        current_row += 1
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_worker_report(data: List[dict], worker: str, period: str, config: dict, for_workers: bool = False) -> bytes:
    """Create individual worker Excel report"""
    worker_normalized = normalize_worker_name(worker.replace(" (оплата клиентом)", ""))
    worker_data = [r for r in data if normalize_worker_name(r.get("worker", "").replace(" (оплата клиентом)", "")) == worker_normalized]
    return create_excel_report(worker_data, period, config, for_workers=for_workers)
