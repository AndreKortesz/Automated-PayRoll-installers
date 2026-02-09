"""
Salary Calculation Service for Montazhniki
FastAPI backend for processing Excel files and calculating salaries

Refactored structure:
- config.py: Configuration and session storage
- utils/: Helper functions (workers, helpers)
- services/: Business logic (geocoding, calculation, excel_parser, excel_report)
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Response
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
from contextlib import asynccontextmanager
from urllib.parse import quote
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import httpx
import json
import os
import re
import zipfile
from io import BytesIO
from datetime import datetime
from typing import Optional, Dict, List, Any
import pathlib

# ============================================================================
# CONFIGURATION (from config.py)
# ============================================================================
from config import DEFAULT_CONFIG, session_data, logger, DEBUG_MODE

# ============================================================================
# DATABASE IMPORTS
# ============================================================================
from database import (
    database, create_tables, connect_db, disconnect_db,
    get_or_create_period, create_upload, save_order, save_calculation,
    save_worker_total, save_change, get_previous_upload, compare_uploads,
    get_orders_by_upload, get_all_periods, get_period_details,
    get_upload_details, get_worker_orders, get_months_summary,
    create_or_update_user, log_action,
    add_duplicate_exclusion, remove_duplicate_exclusion, 
    get_duplicate_exclusions
)

# ============================================================================
# AUTH IMPORTS
# ============================================================================
from auth import (
    get_auth_url, exchange_code_for_token, get_bitrix_user,
    determine_role, is_auth_configured, create_session,
    delete_session, get_current_user, SESSION_COOKIE
)

# ============================================================================
# PERMISSIONS IMPORTS
# ============================================================================
# Note: Permission checks are done inline in endpoints via request.session
# These imports are available for future use if needed:
# from permissions import (
#     check_edit_permission, check_upload_permission,
#     check_delete_row_permission, check_delete_period_permission,
#     check_send_permission, check_send_to_accountant_permission,
#     get_user_permissions, log_user_action, get_client_ip
# )

# ============================================================================
# UTILS IMPORTS (helper functions)
# ============================================================================
from utils import (
    # From utils/helpers.py
    format_order_short,
    format_order_for_workers,
    parse_percent,
    extract_address_from_order,
    clean_address_for_geocoding,
    extract_period,
    # From utils/workers.py
    EXCLUDED_GROUPS,
    build_worker_name_map,
    normalize_worker_name,
    is_valid_worker_name,
)

# ============================================================================
# SERVICES IMPORTS (business logic)
# ============================================================================
from services import (
    # From services/geocoding.py
    geocode_address,
    geocode_address_yandex,
    geocode_address_nominatim,
    get_distance_osrm,
    is_moscow_region,
    calculate_fuel_cost,
    # From services/calculation.py
    calculate_row,
    generate_alarms,
    # From services/excel_parser.py
    parse_excel_file,
    parse_both_excel_files,
    # From services/excel_report.py
    create_excel_report,
    create_worker_report,
)


# ============================================================================
# APPLICATION SETUP
# ============================================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan - startup and shutdown"""
    # Startup
    create_tables()
    await connect_db()
    yield
    # Shutdown
    await disconnect_db()

app = FastAPI(
    title="Salary Calculator", 
    description="Ð Ð°ÑÑ‡Ñ‘Ñ‚ Ð·Ð°Ñ€Ð¿Ð»Ð°Ñ‚Ñ‹ Ð¼Ð¾Ð½Ñ‚Ð°Ð¶Ð½Ð¸ÐºÐ¾Ð²",
    lifespan=lifespan
)

# ============================================================================
# API ROUTERS
# ============================================================================
from api_status import router as status_router
app.include_router(status_router)

# ============================================================================
# CSRF MIDDLEWARE (Optional)
# ============================================================================
# CSRF Protection - uncomment to enable for form-based workflows:
# from csrf_middleware import CSRFMiddleware
# app.add_middleware(CSRFMiddleware)
#
# Currently disabled because:
# 1. API primarily uses JSON requests + session cookies
# 2. File uploads use multipart/form-data with session auth
# Enable if you add traditional HTML form submissions.
# ============================================================================

# Templates and static files
templates = Jinja2Templates(directory="../frontend/templates")

# Create static directory if it doesn't exist (for Railway deployment)
static_dir = pathlib.Path("../frontend/static")
static_dir.mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory="../frontend/static"), name="static")


# ============================================================================
# BUSINESS LOGIC FUNCTIONS
# ============================================================================
# The following functions are now imported from services/ and utils/:
#
# From utils/workers.py:
#   - build_worker_name_map, normalize_worker_name, is_valid_worker_name, EXCLUDED_GROUPS
#
# From utils/helpers.py:
#   - format_order_short, format_order_for_workers, parse_percent
#   - extract_address_from_order, clean_address_for_geocoding, extract_period
#
# From services/geocoding.py:
#   - geocode_address, geocode_address_yandex, geocode_address_nominatim
#   - get_distance_osrm, is_moscow_region, calculate_fuel_cost
#
# From services/calculation.py:
#   - calculate_row, generate_alarms
#
# From services/excel_parser.py:
#   - parse_excel_file, parse_both_excel_files
#
# From services/excel_report.py:
#   - create_excel_report, create_worker_report
# ============================================================================


# ============== AUTH ROUTES ==============

@app.get("/auth/login")
async def auth_login(request: Request):
    """Redirect to Bitrix24 OAuth2 login"""
    if not is_auth_configured():
        # If auth is not configured, redirect to main page (dev mode)
        return RedirectResponse(url="/", status_code=302)

    auth_url = get_auth_url()
    return RedirectResponse(url=auth_url, status_code=302)


@app.get("/auth/callback")
@app.post("/auth/callback")
async def auth_callback(request: Request):
    """Handle Bitrix24 OAuth2 callback (supports both GET and POST)"""

    # Get parameters from query string AND form data
    query_params = dict(request.query_params)
    form_params = {}
    if request.method == "POST":
        form_data = await request.form()
        form_params = dict(form_data)

    # Merge params (form data takes precedence)
    params = {**query_params, **form_params}

    logger.info("ðŸ” Auth callback received: method={request.method}, params={list(params.keys())}")

    # Check for error
    error = params.get("error")
    if error:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð°Ð²Ñ‚Ð¾Ñ€Ð¸Ð·Ð°Ñ†Ð¸Ð¸: {error}",
            "auth_configured": is_auth_configured()
        })

    # Get domain from query params
    domain = params.get("DOMAIN") or params.get("domain") or os.getenv("BITRIX_DOMAIN", "")

    # Bitrix24 local app sends AUTH_ID directly (no code exchange needed!)
    auth_id = params.get("AUTH_ID")
    if auth_id:
        access_token = auth_id
        refresh_token = params.get("REFRESH_ID", "")
        expires_in = int(params.get("AUTH_EXPIRES", 3600))
        logger.info("ðŸ” Using direct AUTH_ID from Bitrix24, expires in {expires_in}s")
    else:
        # Standard OAuth callback with code
        code = params.get("code")
        if not code:
            return templates.TemplateResponse("login.html", {
                "request": request,
                "error": "ÐšÐ¾Ð´ Ð°Ð²Ñ‚Ð¾Ñ€Ð¸Ð·Ð°Ñ†Ð¸Ð¸ Ð½Ðµ Ð¿Ð¾Ð»ÑƒÑ‡ÐµÐ½",
                "auth_configured": is_auth_configured()
            })

        # Exchange code for token (use server_domain from callback params)
        server_domain = params.get("server_domain")
        token_data = await exchange_code_for_token(code, server_domain)
        if not token_data:
            return templates.TemplateResponse("login.html", {
                "request": request,
                "error": "ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ Ð¿Ð¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ñ‚Ð¾ÐºÐµÐ½ Ð´Ð¾ÑÑ‚ÑƒÐ¿Ð°"
            })

        access_token = token_data.get("access_token")
        refresh_token = token_data.get("refresh_token")
        # IMPORTANT: Use domain from callback params, NOT from token_data
        # token_data may contain server_domain (oauth.bitrix24.tech) instead of actual domain
        token_domain = token_data.get("domain")
        if token_domain and "bitrix24.ru" in token_domain:
            domain = token_domain
        # else keep domain from query params (svyaz.bitrix24.ru)
        expires_in = token_data.get("expires_in", 3600)
        
    logger.info("ðŸ” Using domain for user info: {domain}")

    # Get user info from Bitrix24
    bitrix_user = await get_bitrix_user(access_token, domain)
    if not bitrix_user:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ Ð¿Ð¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð¿Ð¾Ð»ÑŒÐ·Ð¾Ð²Ð°Ñ‚ÐµÐ»Ñ"
        })

    # Extract user data
    bitrix_id = int(bitrix_user.get("ID", 0))
    name = f"{bitrix_user.get('LAST_NAME', '')} {bitrix_user.get('NAME', '')}".strip()
    email = bitrix_user.get("EMAIL", "")

    # Determine role
    role = determine_role(bitrix_id)

    # Save/update user in database
    from datetime import timedelta
    token_expires = datetime.utcnow() + timedelta(seconds=expires_in)

    user = await create_or_update_user(
        bitrix_id=bitrix_id,
        name=name,
        email=email,
        role=role,
        access_token=access_token,
        refresh_token=refresh_token,
        token_expires_at=token_expires
    )

    # Create session
    session_user = {
        "id": user["id"],
        "bitrix_id": bitrix_id,
        "name": name,
        "email": email,
        "role": role
    }
    session_id = create_session(session_user)

    # Redirect to main page with session cookie
    response = RedirectResponse(url="/", status_code=302)
    response.set_cookie(
        key=SESSION_COOKIE,
        value=session_id,
        httponly=True,
        max_age=86400,  # 24 hours
        samesite="none",  # Required for iframe context (Bitrix24 app)
        secure=True       # Required when samesite=none
    )

    logger.info("âœ… User logged in: {name} (ID: {bitrix_id}, role: {role})")
    return response


@app.get("/auth/logout")
async def auth_logout(request: Request):
    """Logout user"""
    session_id = request.cookies.get(SESSION_COOKIE)
    if session_id:
        delete_session(session_id)

    response = RedirectResponse(url="/", status_code=302)
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/api/me")
async def get_me(request: Request):
    """Get current user info"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"authenticated": False})

    return JSONResponse({
        "authenticated": True,
        "user": user
    })


@app.get("/login")
async def login_page(request: Request):
    """Show login page"""
    user = get_current_user(request)
    if user:
        return RedirectResponse(url="/", status_code=302)

    return templates.TemplateResponse("login.html", {
        "request": request,
        "auth_configured": is_auth_configured(),
        "auth_url": get_auth_url() if is_auth_configured() else None
    })


# ============== MAIN ROUTES ==============

@app.get("/")
async def index(request: Request):
    user = get_current_user(request)

    # If auth is configured and user is not logged in, redirect to login
    if is_auth_configured() and not user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse("index.html", {
        "request": request,
        "config": DEFAULT_CONFIG,
        "user": user,
        "auth_configured": is_auth_configured()
    })


@app.post("/api/detect-file-type")
async def detect_file_type(file: UploadFile = File(...)):
    """
    Detect if uploaded Excel file contains orders under 10k or over 10k.
    Looks for the filter condition in "ÐžÑ‚Ð±Ð¾Ñ€" row that specifies:
    - "ÐœÐµÐ½ÑŒÑˆÐµ Ð¸Ð»Ð¸ Ñ€Ð°Ð²Ð½Ð¾" -> under 10k
    - "Ð‘Ð¾Ð»ÑŒÑˆÐµ Ð¸Ð»Ð¸ Ñ€Ð°Ð²Ð½Ð¾" -> over 10k
    Also extracts period name from "ÐŸÐµÑ€Ð¸Ð¾Ð´:" row.
    """
    try:
        content = await file.read()
        
        # Parse Excel file
        df = pd.read_excel(BytesIO(content), header=None)
        
        file_type = "unknown"
        period_name = None
        
        # Search through first 20 rows for filter info and period
        for idx in range(min(20, len(df))):
            for col_idx in range(min(10, len(df.columns))):
                cell = df.iloc[idx, col_idx]
                if pd.isna(cell):
                    continue
                cell_str = str(cell).strip()
                
                # Look for filter condition (ÐžÑ‚Ð±Ð¾Ñ€ row)
                # The filter text contains "Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¾Ñ‚ ÑƒÑÐ»ÑƒÐ³ ÐœÐµÐ½ÑŒÑˆÐµ Ð¸Ð»Ð¸ Ñ€Ð°Ð²Ð½Ð¾" or "Ð‘Ð¾Ð»ÑŒÑˆÐµ Ð¸Ð»Ð¸ Ñ€Ð°Ð²Ð½Ð¾"
                if "Ð²Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¾Ñ‚ ÑƒÑÐ»ÑƒÐ³" in cell_str.lower():
                    if "Ð¼ÐµÐ½ÑŒÑˆÐµ Ð¸Ð»Ð¸ Ñ€Ð°Ð²Ð½Ð¾" in cell_str.lower():
                        file_type = "under"
                    elif "Ð±Ð¾Ð»ÑŒÑˆÐµ Ð¸Ð»Ð¸ Ñ€Ð°Ð²Ð½Ð¾" in cell_str.lower():
                        file_type = "over"
                
                # Look for period info
                if "Ð¿ÐµÑ€Ð¸Ð¾Ð´:" in cell_str.lower():
                    # Extract period like "16.11.2025 - 30.11.2025" and normalize to "16-30.11.25"
                    match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})\s*-\s*(\d{2})\.(\d{2})\.(\d{4})', cell_str)
                    if match:
                        d1, m1, y1, d2, m2, y2 = match.groups()
                        period_name = f"{d1}-{d2}.{m1}.{y2[2:]}"
        
        return JSONResponse({
            "success": True, 
            "type": file_type,
            "period": period_name,
            "filename": file.filename
        })
        
    except Exception as e:
        return JSONResponse({
            "success": True,
            "type": "unknown",
            "error": str(e)
        })


@app.post("/upload")
async def upload_files(
    request: Request,
    file_under_10k: UploadFile = File(...),
    file_over_10k: UploadFile = File(...),
    file_yandex_fuel: UploadFile = File(None)  # Optional third file for Yandex Fuel
):
    """Upload and parse Excel files"""
    try:
        content_under = await file_under_10k.read()
        content_over = await file_over_10k.read()
        
        # Parse both files with automatic name normalization
        combined, name_map, manager_comments, parse_warnings = parse_both_excel_files(content_under, content_over)
        
        period_df = pd.read_excel(BytesIO(content_under), header=None)
        period = extract_period(period_df)
        
        # Parse Yandex Fuel file if provided (required for second half periods)
        yandex_fuel_data = {}
        from services.yandex_fuel_parser import parse_yandex_fuel_file, is_second_half_period, validate_yandex_fuel_period
        
        is_second_half = is_second_half_period(period)
        
        if file_yandex_fuel and file_yandex_fuel.filename:
            content_yandex = await file_yandex_fuel.read()
            if content_yandex:
                if is_second_half:
                    # Validate that Yandex file period matches upload period
                    is_valid, error_msg = validate_yandex_fuel_period(content_yandex, period)
                    if not is_valid:
                        return JSONResponse(
                            {"success": False, "detail": f"âŒ ÐÐµÑÐ¾Ð¾Ñ‚Ð²ÐµÑ‚ÑÑ‚Ð²Ð¸Ðµ Ð¿ÐµÑ€Ð¸Ð¾Ð´Ð¾Ð²: {error_msg}"},
                            status_code=400
                        )
                    
                    yandex_fuel_data = parse_yandex_fuel_file(content_yandex, name_map)
                    if DEBUG_MODE: logger.debug("â›½ Ð¯Ð½Ð´ÐµÐºÑ Ð—Ð°Ð¿Ñ€Ð°Ð²ÐºÐ¸ Ð·Ð°Ð³Ñ€ÑƒÐ¶ÐµÐ½Ñ‹: {len(yandex_fuel_data)} Ð¼Ð¾Ð½Ñ‚Ð°Ð¶Ð½Ð¸ÐºÐ¾Ð²")
                else:
                    logger.warning("âš ï¸ Ð¯Ð½Ð´ÐµÐºÑ Ð—Ð°Ð¿Ñ€Ð°Ð²ÐºÐ¸: Ð¿ÐµÑ€Ð¸Ð¾Ð´ {period} - Ð¿ÐµÑ€Ð²Ð°Ñ Ð¿Ð¾Ð»Ð¾Ð²Ð¸Ð½Ð° Ð¼ÐµÑÑÑ†Ð°, Ñ„Ð°Ð¹Ð» Ð¸Ð³Ð½Ð¾Ñ€Ð¸Ñ€ÑƒÐµÑ‚ÑÑ")
        elif is_second_half:
            # Yandex Fuel file is required for second half periods
            return JSONResponse(
                {"success": False, "detail": "Ð”Ð»Ñ Ð²Ñ‚Ð¾Ñ€Ð¾Ð¹ Ð¿Ð¾Ð»Ð¾Ð²Ð¸Ð½Ñ‹ Ð¼ÐµÑÑÑ†Ð° Ð½ÐµÐ¾Ð±Ñ…Ð¾Ð´Ð¸Ð¼Ð¾ Ð·Ð°Ð³Ñ€ÑƒÐ·Ð¸Ñ‚ÑŒ Ñ„Ð°Ð¹Ð» Ð¯Ð½Ð´ÐµÐºÑ Ð—Ð°Ð¿Ñ€Ð°Ð²Ð¾Ðº"},
                status_code=400
            )
        
        workers = list(set([w.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "") 
                           for w in combined["worker"].unique() if w and not pd.isna(w)]))
        workers = sorted(workers)
        
        orders = []
        for _, row in combined.iterrows():
            order = row.get("order", "")
            worker = row.get("worker", "")
            is_client = row.get("is_client_payment", False)
            
            # Get revenue_services and percent for transport check
            revenue_services = row.get("revenue_services", 0)
            revenue_services = float(revenue_services) if pd.notna(revenue_services) and revenue_services != "" else 0
            
            percent_str = str(row.get("percent", "0%")).replace("%", "").replace(",", ".").strip()
            try:
                percent = float(percent_str) if percent_str else 0
            except (ValueError, TypeError):
                percent = 0
            
            # Check if transport will be applied (revenue > 10k and percent between 20% and 40%)
            percent_min = DEFAULT_CONFIG["transport_percent_min"]
            percent_max = DEFAULT_CONFIG["transport_percent_max"]
            has_transport = revenue_services > DEFAULT_CONFIG["transport_min_revenue"] and percent_min <= percent <= percent_max
            
            if order and not str(order).startswith(("ÐžÐ‘Ð£Ð§Ð•ÐÐ˜Ð•", "Ð’ Ð¿Ñ€Ð¾ÑˆÐ»Ð¾Ð¼")):
                orders.append({
                    "worker": worker.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""),
                    "order": order,
                    "order_short": format_order_short(order),
                    "is_client_payment": is_client,
                    "has_transport": has_transport
                })
        
        orders.sort(key=lambda x: x["worker"])
        
        session_id = datetime.now().strftime("%Y%m%d%H%M%S")
        session_data[session_id] = {
            "combined": combined.to_dict("records"),
            "period": period,
            "workers": workers,
            "name_map": name_map,  # Save for later use
            "yandex_fuel": yandex_fuel_data,  # Yandex Fuel deductions per worker
            "manager_comments": manager_comments,  # Manager comments for orders
            "parse_warnings": parse_warnings  # Warnings (e.g., managers in data)
        }
        
        # ===== CHECK FOR CHANGES FROM PREVIOUS UPLOAD =====
        changes_summary = {
            "has_previous": False,
            "added": [],
            "deleted": [],
            "modified": []
        }
        
        try:
            if database and database.is_connected:
                from database import get_or_create_period, get_period_details, get_orders_by_upload
                
                # Check if this period exists
                period_id = await get_or_create_period(period)
                period_details = await get_period_details(period_id)
                
                if period_details and period_details.get("uploads"):
                    uploads = period_details["uploads"]
                    if len(uploads) > 0:
                        # Find latest upload with actual orders (skip empty uploads)
                        latest_upload_id = None
                        latest_upload_version = None
                        latest_upload_date = None
                        old_orders = []
                        
                        for upload in uploads:
                            upload_id_check = upload["id"]
                            orders_check = await get_orders_by_upload(upload_id_check)
                            if orders_check and len(orders_check) > 0:
                                latest_upload_id = upload_id_check
                                latest_upload_version = upload["version"]
                                latest_upload_date = str(upload.get("created_at", ""))
                                old_orders = orders_check
                                if DEBUG_MODE: logger.debug("ðŸ“Š Found version {latest_upload_version} with {len(old_orders)} orders")
                                break
                            else:
                                logger.warning("âš ï¸ Skipping empty version {upload['version']}")
                        
                        if not latest_upload_id:
                            if DEBUG_MODE: logger.debug("ðŸ“Š No previous version with orders found")
                        
                        # Also get extra rows (manual additions) from previous upload
                        from database import get_upload_details
                        prev_upload_details = await get_upload_details(latest_upload_id) if latest_upload_id else None
                        extra_rows_from_prev = []
                        manual_edits_from_prev = []
                        
                        if prev_upload_details:
                            # Get extra rows (is_extra_row=True)
                            for o in old_orders:
                                is_extra = o.get("is_extra_row", False)
                                if is_extra:
                                    extra_rows_from_prev.append(o)
                                    if DEBUG_MODE: logger.debug("ðŸ“‹ Found extra_row: {o.get('order_code', '')} - {o.get('worker', '')}")
                            
                            if DEBUG_MODE: logger.debug("ðŸ“‹ Total extra_rows found: {len(extra_rows_from_prev)} out of {len(old_orders)} orders")
                            
                            # Debug: show is_extra_row values
                            extra_counts = {"True": 0, "False": 0, "None": 0}
                            for o in old_orders:
                                val = o.get("is_extra_row")
                                if val is True:
                                    extra_counts["True"] += 1
                                elif val is False:
                                    extra_counts["False"] += 1
                                else:
                                    extra_counts["None"] += 1
                            if DEBUG_MODE: logger.debug("ðŸ“‹ is_extra_row distribution: {extra_counts}")
                            
                            # Get manual edits
                            manual_edits_from_prev = prev_upload_details.get("manual_edits", [])
                        
                        if old_orders:
                            changes_summary["has_previous"] = True
                            changes_summary["previous_version"] = latest_upload_version
                            changes_summary["previous_date"] = latest_upload_date
                            changes_summary["previous_upload_id"] = latest_upload_id
                            
                            # Build maps for comparison
                            # IMPORTANT: Normalize worker names from DB using current name_map
                            # to ensure consistent keys between old and new data
                            old_map = {}
                            for o in old_orders:
                                # Skip extra rows for normal comparison
                                if o.get("is_extra_row", False):
                                    continue
                                # Normalize worker name from DB to match new file names
                                worker_from_db = o.get("worker", "")
                                worker_normalized = normalize_worker_name(worker_from_db, name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
                                key = (o.get("order_code", ""), worker_normalized)
                                old_map[key] = o
                            
                            if DEBUG_MODE: logger.debug("ðŸ“Š Comparison: {len(old_map)} orders in DB")
                            
                            new_map = {}
                            for _, row in combined.iterrows():
                                # Skip worker total rows
                                if row.get("is_worker_total", False):
                                    continue
                                    
                                order_text = str(row.get("order", ""))
                                order_code_match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                                order_code = order_code_match.group(0) if order_code_match else ""
                                
                                # Skip rows without order code (they are totals or headers)
                                if not order_code:
                                    continue
                                    
                                # IMPORTANT: Use name_map for consistent normalization with old_map
                                worker = normalize_worker_name(str(row.get("worker", "")), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
                                
                                # Extract address from order text using proper function
                                address = extract_address_from_order(order_text)
                                if not address and ", " in order_text:
                                    # Fallback: simple extraction
                                    parts = order_text.split(", ", 1)
                                    if len(parts) > 1:
                                        address = parts[1].split("\n")[0][:80]
                                
                                key = (order_code, worker)
                                
                                # Collect all numeric fields - with safe parsing
                                def safe_float(val):
                                    if val is None or val == "" or pd.isna(val):
                                        return 0.0
                                    try:
                                        if isinstance(val, str):
                                            val = val.replace(" ", "").replace(",", ".").replace("%", "")
                                        return float(val)
                                    except (ValueError, TypeError):
                                        return 0.0
                                
                                revenue_total = safe_float(row.get("revenue_total", 0))
                                revenue_services = safe_float(row.get("revenue_services", 0))
                                diagnostic = safe_float(row.get("diagnostic", 0))
                                specialist_fee = safe_float(row.get("specialist_fee", 0))
                                additional_expenses = safe_float(row.get("additional_expenses", 0))
                                service_payment = safe_float(row.get("service_payment", 0))
                                percent_val = parse_percent(row.get("percent", 0))
                                
                                new_map[key] = {
                                    "order_code": order_code,
                                    "order_full": order_text,
                                    "address": address,
                                    "worker": worker,
                                    "revenue_total": revenue_total,
                                    "revenue_services": revenue_services,
                                    "diagnostic": diagnostic,
                                    "specialist_fee": specialist_fee,
                                    "additional_expenses": additional_expenses,
                                    "service_payment": service_payment,
                                    "percent": percent_val,
                                }
                            
                            if DEBUG_MODE: logger.debug("ðŸ“Š Comparison: {len(new_map)} orders in new files")
                            
                            # Calculate fuel and transport for new orders BEFORE comparison
                            # This ensures we compare apples to apples
                            config = DEFAULT_CONFIG.copy()
                            company_car_workers = config.get("company_car_workers", [])
                            company_car_normalized = [normalize_worker_name(w) for w in company_car_workers]
                            
                            for key, order in new_map.items():
                                # Calculate fuel
                                fuel_payment = 0
                                if order["specialist_fee"] == 0 and order["address"]:
                                    fuel_payment = await calculate_fuel_cost(order["address"], config, 1)
                                order["fuel_payment"] = fuel_payment
                                
                                # Calculate transport
                                transport = 0
                                worker_normalized = normalize_worker_name(order["worker"])
                                is_on_company_car = worker_normalized in company_car_normalized
                                percent_min = config.get("transport_percent_min", 20)
                                percent_max = config.get("transport_percent_max", 40)
                                if order["revenue_services"] > config["transport_min_revenue"] and percent_min <= order["percent"] <= percent_max:
                                    if not is_on_company_car:
                                        transport = config["transport_amount"]
                                order["transport"] = transport
                                
                                # Calculate total
                                order["total"] = order["service_payment"] + fuel_payment + transport
                            
                            if DEBUG_MODE: logger.debug("ðŸ“Š Calculated fuel/transport for {len(new_map)} new orders")
                            
                            # Find added - include all details
                            for key, order in new_map.items():
                                if key[0] and key not in old_map:  # Has order_code and not in old
                                    # Build details dict with non-zero values
                                    details = {}
                                    if order["revenue_total"] > 0:
                                        details["Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¸Ñ‚Ð¾Ð³Ð¾"] = f"{order['revenue_total']:,.0f}".replace(",", " ")
                                    if order["revenue_services"] != 0:
                                        details["Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¾Ñ‚ ÑƒÑÐ»ÑƒÐ³"] = f"{order['revenue_services']:,.0f}".replace(",", " ")
                                    if order["diagnostic"] > 0:
                                        details["Ð”Ð¸Ð°Ð³Ð½Ð¾ÑÑ‚Ð¸ÐºÐ°"] = f"{order['diagnostic']:,.0f}".replace(",", " ")
                                    if order["specialist_fee"] > 0:
                                        details["Ð’Ñ‹ÐµÐ·Ð´ ÑÐ¿ÐµÑ†Ð¸Ð°Ð»Ð¸ÑÑ‚Ð°"] = f"{order['specialist_fee']:,.0f}".replace(",", " ")
                                    if order["additional_expenses"] != 0:
                                        details["Ð”Ð¾Ð¿. Ñ€Ð°ÑÑ…Ð¾Ð´Ñ‹"] = f"{order['additional_expenses']:,.0f}".replace(",", " ")
                                    if order["service_payment"] != 0:
                                        details["ÐžÐ¿Ð»Ð°Ñ‚Ð° ÑƒÑÐ»ÑƒÐ³"] = f"{order['service_payment']:,.0f}".replace(",", " ")
                                    if order["percent"] > 0:
                                        details["ÐŸÑ€Ð¾Ñ†ÐµÐ½Ñ‚"] = f"{order['percent']:.0f}%"
                                    
                                    changes_summary["added"].append({
                                        "order_code": order["order_code"],
                                        "worker": order["worker"],
                                        "address": order["address"],
                                        "details": details
                                    })
                            
                            # Find deleted - include address from old data
                            def safe_float_db(val):
                                """Safely parse float from DB value that might be string like '30,00 %'"""
                                if val is None or val == "":
                                    return 0.0
                                try:
                                    if isinstance(val, (int, float)):
                                        return float(val)
                                    val_str = str(val).replace(" ", "").replace(",", ".").replace("%", "")
                                    return float(val_str)
                                except (ValueError, TypeError):
                                    return 0.0
                            
                            for key, order in old_map.items():
                                if key[0] and key not in new_map:  # Has order_code and not in new
                                    # Build details from old order
                                    details = {}
                                    if safe_float_db(order.get("revenue_total", 0)) > 0:
                                        details["Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¸Ñ‚Ð¾Ð³Ð¾"] = f"{safe_float_db(order.get('revenue_total', 0)):,.0f}".replace(",", " ")
                                    if safe_float_db(order.get("revenue_services", 0)) != 0:
                                        details["Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¾Ñ‚ ÑƒÑÐ»ÑƒÐ³"] = f"{safe_float_db(order.get('revenue_services', 0)):,.0f}".replace(",", " ")
                                    if safe_float_db(order.get("service_payment", 0)) != 0:
                                        details["ÐžÐ¿Ð»Ð°Ñ‚Ð° ÑƒÑÐ»ÑƒÐ³"] = f"{safe_float_db(order.get('service_payment', 0)):,.0f}".replace(",", " ")
                                    if safe_float_db(order.get("percent", 0)) > 0:
                                        details["ÐŸÑ€Ð¾Ñ†ÐµÐ½Ñ‚"] = f"{safe_float_db(order.get('percent', 0)):.0f}%"
                                    
                                    changes_summary["deleted"].append({
                                        "order_code": order.get("order_code", ""),
                                        "worker": order.get("worker", ""),
                                        "address": order.get("address", ""),
                                        "details": details
                                    })
                            
                            # Log extra rows (they will be added to deleted later with full details)
                            for extra_order in extra_rows_from_prev:
                                # total comes directly from JOIN query now
                                calc_total = extra_order.get("total", 0) or 0
                                order_text = extra_order.get("order", "") or extra_order.get("order_full", "")
                                if DEBUG_MODE: logger.debug("ðŸ“‹ Found extra row: {order_text[:30]}_{extra_order.get('worker', '')} total={calc_total}")
                            
                            # Find modified - compare all fields
                            compare_fields = [
                                ("revenue_total", "Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¸Ñ‚Ð¾Ð³Ð¾"),
                                ("revenue_services", "Ð’Ñ‹Ñ€ÑƒÑ‡ÐºÐ° Ð¾Ñ‚ ÑƒÑÐ»ÑƒÐ³"),
                                ("diagnostic", "Ð”Ð¸Ð°Ð³Ð½Ð¾ÑÑ‚Ð¸ÐºÐ°"),
                                ("specialist_fee", "Ð’Ñ‹ÐµÐ·Ð´ ÑÐ¿ÐµÑ†Ð¸Ð°Ð»Ð¸ÑÑ‚Ð°"),
                                ("additional_expenses", "Ð”Ð¾Ð¿. Ñ€Ð°ÑÑ…Ð¾Ð´Ñ‹"),
                                ("service_payment", "ÐžÐ¿Ð»Ð°Ñ‚Ð° ÑƒÑÐ»ÑƒÐ³"),
                                ("percent", "ÐŸÑ€Ð¾Ñ†ÐµÐ½Ñ‚"),
                            ]
                            
                            # Debug: show some keys from both maps
                            if DEBUG_MODE: logger.debug("ðŸ“Š Sample old_map keys: {list(old_map.keys())[:5]}")
                            if DEBUG_MODE: logger.debug("ðŸ“Š Sample new_map keys: {list(new_map.keys())[:5]}")

                            # Debug: find ÐšÐÐ£Ð¢-001143 specifically - compare SAME worker in both maps
                            debug_order = "ÐšÐÐ£Ð¢-001143"
                            old_keys_with_debug = [k for k in old_map.keys() if debug_order in k[0]]
                            new_keys_with_debug = [k for k in new_map.keys() if debug_order in k[0]]
                            if DEBUG_MODE: logger.debug("ðŸ” {debug_order} in old_map: {old_keys_with_debug}")
                            if DEBUG_MODE: logger.debug("ðŸ” {debug_order} in new_map: {new_keys_with_debug}")

                            # Compare EACH worker for this order between old and new
                            for key in old_keys_with_debug:
                                if key in new_map:
                                    old_data = old_map[key]
                                    new_data = new_map[key]
                                    if DEBUG_MODE: logger.debug("ðŸ” comparing {key}:")
                                    if DEBUG_MODE: logger.debug(f"   OLD: rt={old_data.get('revenue_total')}, rs={old_data.get('revenue_services')}, sp={old_data.get('service_payment')}")
                                    if DEBUG_MODE: logger.debug(f"   NEW: rt={new_data.get('revenue_total')}, rs={new_data.get('revenue_services')}, sp={new_data.get('service_payment')}")
                                    # Check differences
                                    for field in ['revenue_total', 'revenue_services', 'service_payment', 'diagnostic', 'specialist_fee']:
                                        old_val = safe_float_db(old_data.get(field, 0))
                                        new_val = float(new_data.get(field, 0) or 0)
                                        if abs(old_val - new_val) > 0.01:
                                            if DEBUG_MODE: logger.debug(f"   âš ï¸ DIFF {field}: {old_val} â†’ {new_val}")
                                else:
                                    if DEBUG_MODE: logger.debug("ðŸ” {key} NOT in new_map - will be DELETED")

                            # Debug: compare a sample order
                            for key in list(new_map.keys())[:3]:
                                if key in old_map:
                                    old_o = old_map[key]
                                    new_o = new_map[key]
                                    if DEBUG_MODE: logger.debug("ðŸ“Š Sample compare {key}:")
                                    if DEBUG_MODE: logger.debug(f"   old revenue_total={old_o.get('revenue_total')} ({type(old_o.get('revenue_total')).__name__})")
                                    if DEBUG_MODE: logger.debug(f"   new revenue_total={new_o.get('revenue_total')} ({type(new_o.get('revenue_total')).__name__})")
                            
                            for key in new_map:
                                if key[0] and key in old_map:  # Both exist
                                    old_order = old_map[key]
                                    new_order = new_map[key]

                                    field_changes = []
                                    for field_key, field_name in compare_fields:
                                        # old_order comes from DB - might have string values like '30,00 %'
                                        old_val = safe_float_db(old_order.get(field_key, 0))
                                        # new_order comes from parsed file - already numeric
                                        new_val = float(new_order.get(field_key, 0) or 0)

                                        if abs(old_val - new_val) > 0.01:  # Compare with tolerance
                                            if field_key == "percent":
                                                field_changes.append({
                                                    "field": field_name,
                                                    "old": f"{old_val:.0f}%",
                                                    "new": f"{new_val:.0f}%"
                                                })
                                            else:
                                                field_changes.append({
                                                    "field": field_name,
                                                    "old": f"{old_val:,.0f}".replace(",", " "),
                                                    "new": f"{new_val:,.0f}".replace(",", " ")
                                                })

                                    # IMPORTANT: Compare calculated totals (new has fuel/transport already calculated)
                                    # This detects manual edits made in UI
                                    old_total = safe_float_db(old_order.get("total", 0))
                                    old_fuel = safe_float_db(old_order.get("fuel_payment", 0))
                                    old_transport = safe_float_db(old_order.get("transport", 0))
                                    
                                    # new_order now has calculated fuel/transport/total
                                    new_total = float(new_order.get("total", 0) or 0)
                                    new_fuel = float(new_order.get("fuel_payment", 0) or 0)
                                    new_transport = float(new_order.get("transport", 0) or 0)

                                    # Calculate what total SHOULD be with old fuel/transport
                                    # This isolates manual edits from fuel calculation fluctuations
                                    new_service_payment = float(new_order.get("service_payment", 0) or 0)
                                    expected_total_with_old_fuel = new_service_payment + old_fuel + old_transport
                                    
                                    # Compare old_total with expected - if different, there was a REAL manual edit
                                    # (not just fuel API fluctuation)
                                    if abs(old_total - expected_total_with_old_fuel) > 0.01:
                                        field_changes.append({
                                            "field": "Ð˜Ñ‚Ð¾Ð³Ð¾ (Ñ€ÑƒÑ‡Ð½Ð¾Ðµ Ð¸Ð·Ð¼ÐµÐ½ÐµÐ½Ð¸Ðµ)",
                                            "old": f"{old_total:,.0f}".replace(",", " "),
                                            "new": f"{expected_total_with_old_fuel:,.0f}".replace(",", " ") + " (Ð¿ÐµÑ€ÐµÑÑ‡Ð¸Ñ‚Ð°Ð½Ð¾)"
                                        })
                                        if DEBUG_MODE: logger.debug("ðŸ“Š Manual edit detected: {key} - old_total={old_total}, expected={expected_total_with_old_fuel}")
                                    
                                    # Check if transport differs (this is a real change, not API fluctuation)
                                    if abs(old_transport - new_transport) > 0.01:
                                        field_changes.append({
                                            "field": "Ð¢Ñ€Ð°Ð½ÑÐ¿Ð¾Ñ€Ñ‚Ð½Ñ‹Ðµ",
                                            "old": f"{old_transport:,.0f}".replace(",", " "),
                                            "new": f"{new_transport:,.0f}".replace(",", " ")
                                        })
                                        if DEBUG_MODE: logger.debug("ðŸ“Š Transport differs: {key} - old={old_transport}, new={new_transport}")
                                    
                                    # Check if fuel differs SIGNIFICANTLY (more than API fluctuation)
                                    if abs(old_fuel - new_fuel) > 250:
                                        field_changes.append({
                                            "field": "Ð‘ÐµÐ½Ð·Ð¸Ð½",
                                            "old": f"{old_fuel:,.0f}".replace(",", " "),
                                            "new": f"{new_fuel:,.0f}".replace(",", " ")
                                        })
                                        if DEBUG_MODE: logger.debug("ðŸ“Š Fuel differs significantly: {key} - old={old_fuel}, new={new_fuel}")

                                    if field_changes:
                                        if DEBUG_MODE: logger.debug("ðŸ“Š Modified found: {key} - {field_changes}")
                                        changes_summary["modified"].append({
                                            "order_code": new_order["order_code"],
                                            "worker": new_order["worker"],
                                            "address": new_order["address"],
                                            "changes": field_changes
                                        })
                            
                            if DEBUG_MODE: logger.debug("ðŸ“Š Comparison result: {len(changes_summary['added'])} added, {len(changes_summary['deleted'])} deleted, {len(changes_summary['modified'])} modified")
                            
                            # Add extra rows (manual additions) from previous version to deleted list
                            # These are rows that were manually added and won't be in new 1C files
                            if DEBUG_MODE: logger.debug("ðŸ“‹ DEBUG: extra_rows_from_prev has {len(extra_rows_from_prev)} items before loop")
                            changes_summary["extra_rows"] = []  # Store for later restoration
                            for extra in extra_rows_from_prev:
                                order_text = extra.get("order_full", "") or extra.get("order", "") or extra.get("order_code", "")
                                worker = extra.get("worker", "")
                                
                                if DEBUG_MODE: logger.debug("ðŸ“‹ Extra row from DB: order_full='{extra.get('order_full', '')}', order='{extra.get('order', '')}', order_code='{extra.get('order_code', '')}' -> order_text='{order_text}'")
                                
                                # Get total directly from extra (from JOIN query)
                                total = extra.get("total", 0) or 0
                                
                                changes_summary["extra_rows"].append({
                                    "id": extra.get("id"),
                                    "order_code": extra.get("order_code", "") or order_text[:50],
                                    "order_full": order_text,
                                    "worker": worker,
                                    "address": extra.get("address", ""),
                                    "total": total,
                                    "type": "extra_row"
                                })
                                
                                # Add to deleted list for UI
                                changes_summary["deleted"].append({
                                    "order_code": order_text[:50] if not extra.get("order_code") else extra.get("order_code"),
                                    "worker": worker,
                                    "address": extra.get("address", "") or order_text,
                                    "details": {
                                        "Ð˜Ñ‚Ð¾Ð³Ð¾": f"{total:,.0f}".replace(",", " ") if total else "â€”"
                                    },
                                    "type": "extra_row",
                                    "original_id": extra.get("id")
                                })
                            
                            # Also add manual_edits info for potential restoration
                            changes_summary["manual_edits_prev"] = manual_edits_from_prev
                            
                            if DEBUG_MODE: logger.debug("ðŸ“‹ FINAL: changes_summary has {len(changes_summary['added'])} added, {len(changes_summary['deleted'])} deleted (including {len(changes_summary.get('extra_rows', []))} extra_rows)")
                            
        except Exception as e:
            logger.warning("âš ï¸ Changes comparison error (non-critical): {e}")
            import traceback
            traceback.print_exc()
        
        # Store changes_summary in session for review page
        session_data[session_id]["changes_summary"] = changes_summary
        
        # Check if there are changes to review
        has_changes = (
            changes_summary.get("has_previous", False) and (
                len(changes_summary.get("added", [])) > 0 or
                len(changes_summary.get("deleted", [])) > 0 or
                len(changes_summary.get("modified", [])) > 0
            )
        )
        
        # Get manager comments and warnings from session
        manager_comments = session_data[session_id].get("manager_comments", [])
        parse_warnings = session_data[session_id].get("parse_warnings", [])
        has_manager_comments = len(manager_comments) > 0
        has_warnings = len(parse_warnings) > 0
        
        return JSONResponse({
            "success": True,
            "session_id": session_id,
            "period": period,
            "workers": workers,
            "orders": orders,
            "total_records": len(combined),
            "changes": changes_summary,
            "has_changes": has_changes,
            "redirect_to_review": has_changes or has_manager_comments or has_warnings,  # Also redirect if has warnings
            "manager_comments": manager_comments,
            "has_manager_comments": has_manager_comments,
            "parse_warnings": parse_warnings,
            "has_warnings": has_warnings
        })
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ===== REVIEW CHANGES PAGE =====
@app.get("/review")
async def review_page(request: Request):
    """Render the review changes page"""
    user = get_current_user(request)
    return templates.TemplateResponse("review.html", {"request": request, "user": user})


@app.get("/api/review/{session_id}")
async def get_review_data(session_id: str):
    """Get changes data for review page"""
    try:
        if session_id not in session_data:
            return JSONResponse({"success": False, "error": "Ð¡ÐµÑÑÐ¸Ñ Ð¸ÑÑ‚ÐµÐºÐ»Ð°"})
        
        session = session_data[session_id]
        changes = session.get("changes_summary", {})
        manager_comments = session.get("manager_comments", [])
        parse_warnings = session.get("parse_warnings", [])
        
        return JSONResponse({
            "success": True,
            "session_id": session_id,
            "period": session.get("period", ""),
            "previous_version": changes.get("previous_version", 1),
            "changes": {
                "added": changes.get("added", []),
                "deleted": changes.get("deleted", []),
                "modified": changes.get("modified", [])
            },
            "manager_comments": manager_comments,
            "parse_warnings": parse_warnings
        })
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})


@app.post("/api/apply-review")
async def apply_review_changes(request: Request):
    """Apply selected changes and proceed with calculation"""
    try:
        data = await request.json()
        session_id = data.get("session_id")
        selections = data.get("selections", {})
        
        if session_id not in session_data:
            return JSONResponse({"success": False, "error": "Ð¡ÐµÑÑÐ¸Ñ Ð¸ÑÑ‚ÐµÐºÐ»Ð°"})
        
        session = session_data[session_id]
        combined_records = session.get("combined", [])
        changes = session.get("changes_summary", {})
        name_map = session.get("name_map", {})  # Get name_map for consistent normalization
        
        # Convert combined back to list of dicts for modification
        modified_records = list(combined_records)
        
        # Process deleted items that should be restored
        deleted_to_restore = selections.get("deleted", [])
        if deleted_to_restore and changes.get("has_previous"):
            # We need to fetch the old orders from database
            try:
                from database import get_orders_by_upload, get_upload_details
                period_id = await get_or_create_period(session["period"])
                period_details = await get_period_details(period_id)
                
                if period_details and period_details.get("uploads"):
                    # Find latest upload with actual orders (skip empty uploads)
                    latest_upload_id = None
                    old_orders = []
                    for upload in period_details["uploads"]:
                        upload_id_check = upload["id"]
                        orders_check = await get_orders_by_upload(upload_id_check)
                        if orders_check and len(orders_check) > 0:
                            latest_upload_id = upload_id_check
                            old_orders = orders_check
                            if DEBUG_MODE: logger.debug("ðŸ“‹ Found previous version with {len(old_orders)} orders for restoration")
                            break
                    
                    if not latest_upload_id or not old_orders:
                        logger.warning("âš ï¸ No previous version with orders found for restoration")
                    else:
                        # Get upload details for extra rows with calculations
                        upload_details = await get_upload_details(latest_upload_id)
                        
                        for old_order in old_orders:
                            order_code = old_order.get("order_code", "")
                            order_full = old_order.get("order_full", "")
                            worker = old_order.get("worker", "")
                            is_extra = old_order.get("is_extra_row", False)
                            
                            # For extra rows, use order_full as key if no order_code
                            if is_extra and not order_code:
                                key = (order_full[:50] if order_full else "EXTRA") + "_" + worker
                            else:
                                key = order_code + "_" + worker
                            
                            if key in deleted_to_restore:
                                # Get calculation data directly from old_order (from JOIN query)
                                calc_total = old_order.get("total", 0) or 0
                                calc_fuel = old_order.get("fuel_payment", 0) or 0
                                calc_transport = old_order.get("transport", 0) or 0
                                
                                if DEBUG_MODE: logger.debug("ðŸ“‹ Restoring {key}: total={calc_total}, fuel={calc_fuel}, transport={calc_transport}")
                                
                                # Add this order back to combined records
                                restored_record = {
                                    "worker": worker,
                                    "order": order_full or order_code,
                                    "order_code": order_code,  # Preserve original order_code for extra rows
                                    "address": old_order.get("address", ""),  # Preserve address
                                    "revenue_total": old_order.get("revenue_total", 0),
                                    "revenue_services": old_order.get("revenue_services", 0),
                                    "diagnostic": old_order.get("diagnostic", 0),
                                    "diagnostic_payment": old_order.get("diagnostic_payment", 0),
                                    "specialist_fee": old_order.get("specialist_fee", 0),
                                    "additional_expenses": old_order.get("additional_expenses", 0),
                                    "service_payment": old_order.get("service_payment", 0),
                                    "percent": old_order.get("percent", 0),
                                    "is_client_payment": old_order.get("is_client_payment", False),
                                    "is_restored": True,  # Mark as restored
                                    "is_extra_row": is_extra,
                                    # Preserve calculation values for extra rows
                                    "fuel_payment": calc_fuel,
                                    "transport": calc_transport,
                                    "total": calc_total,
                                }
                                modified_records.append(restored_record)
                                logger.info("âœ… Restored: {key} (extra_row={is_extra}, total={calc_total})")
            except Exception as e:
                logger.error(f"Error restoring deleted orders: {e}")
                import traceback
                traceback.print_exc()
        
        # Process modified items where user wants to keep old values
        modified_to_revert = selections.get("modified", [])
        if modified_to_revert and changes.get("has_previous"):
            try:
                from database import get_orders_by_upload
                period_id = await get_or_create_period(session["period"])
                period_details = await get_period_details(period_id)
                
                if period_details and period_details.get("uploads"):
                    # Find latest upload with actual orders
                    latest_upload_id = None
                    old_orders = []
                    for upload in period_details["uploads"]:
                        upload_id_check = upload["id"]
                        orders_check = await get_orders_by_upload(upload_id_check)
                        if orders_check and len(orders_check) > 0:
                            latest_upload_id = upload_id_check
                            old_orders = orders_check
                            break
                    
                    if old_orders:
                        old_orders_map = {}
                        for o in old_orders:
                            # Normalize worker name for consistent key matching
                            worker_normalized = normalize_worker_name(o.get("worker", ""), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
                            old_orders_map[o.get("order_code", "") + "_" + worker_normalized] = o

                        # Update records with old values
                        for i, record in enumerate(modified_records):
                            order_text = str(record.get("order", ""))
                            order_code_match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                            order_code = order_code_match.group(0) if order_code_match else ""
                            # Normalize worker name for consistent key matching
                            worker = normalize_worker_name(str(record.get("worker", "")), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
                            key = order_code + "_" + worker
                            
                            if key in modified_to_revert and key in old_orders_map:
                                old = old_orders_map[key]
                                # Revert numeric fields to old values
                                modified_records[i]["revenue_total"] = old.get("revenue_total", 0)
                                modified_records[i]["revenue_services"] = old.get("revenue_services", 0)
                                modified_records[i]["diagnostic"] = old.get("diagnostic", 0)
                                modified_records[i]["specialist_fee"] = old.get("specialist_fee", 0)
                                modified_records[i]["additional_expenses"] = old.get("additional_expenses", 0)
                                modified_records[i]["service_payment"] = old.get("service_payment", 0)
                                modified_records[i]["percent"] = old.get("percent", 0)
                                modified_records[i]["is_reverted"] = True

                                # IMPORTANT: Also preserve calculation values (including manual edits)
                                # This ensures "Ð’Ð°Ñ€Ð¸Ð°Ð½Ñ‚ B" works - old version values are kept
                                # Data comes directly from JOIN query now
                                old_total = old.get("total", 0) or 0
                                old_fuel = old.get("fuel_payment", 0) or 0
                                old_transport = old.get("transport", 0) or 0
                                if old_total or old_fuel or old_transport:
                                    modified_records[i]["_old_calc_total"] = old_total
                                    modified_records[i]["_old_calc_fuel"] = old_fuel
                                    modified_records[i]["_old_calc_transport"] = old_transport
                                    if DEBUG_MODE: logger.debug("ðŸ“‹ Preserving old calc for {key}: total={old_total}")
            except Exception as e:
                logger.error(f"Error reverting modified orders: {e}")
        
        # Process added items - selections.added contains keys to SKIP (not add)
        added_to_skip = set(selections.get("added", []))
        
        if added_to_skip:
            if DEBUG_MODE: logger.debug("ðŸ“‹ Skipping {len(added_to_skip)} added orders: {added_to_skip}")
            filtered_records = []
            for record in modified_records:
                order_text = str(record.get("order", ""))
                order_code_match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                order_code = order_code_match.group(0) if order_code_match else ""
                # Normalize worker name for consistent key matching
                worker = normalize_worker_name(str(record.get("worker", "")), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
                key = order_code + "_" + worker

                if key not in added_to_skip:
                    filtered_records.append(record)
                else:
                    if DEBUG_MODE: logger.debug(f"   â­ï¸ Skipping: {key}")
            modified_records = filtered_records
        
        # Update session with modified records
        session["combined"] = modified_records
        session["review_applied"] = True
        session_data[session_id] = session
        
        # Process manager comment selections
        # manager_comment_selections is a dict: {order_key: true/false}
        # where true = apply manager's payment, false = use standard calculation
        manager_selections = selections.get("manager_comments", {})
        applied_manager_comments = {}  # Track which orders have manager overrides
        
        if manager_selections:
            for record in modified_records:
                order_text = str(record.get("order", ""))
                order_code_match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                order_code = order_code_match.group(0) if order_code_match else ""
                worker = normalize_worker_name(str(record.get("worker", "")), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
                key = order_code + "_" + worker
                
                if key in manager_selections and manager_selections[key]:
                    # Apply manager comment
                    parsed = record.get("manager_comment_parsed")
                    if parsed:
                        if parsed["type"] == "percent":
                            # Calculate new service_payment based on percent
                            revenue_services = float(record.get("revenue_services", 0) or 0)
                            new_payment = revenue_services * (parsed["value"] / 100)
                            record["service_payment"] = new_payment
                            record["manager_override"] = True
                            record["manager_override_type"] = "percent"
                            record["manager_override_value"] = parsed["value"]
                            applied_manager_comments[key] = {
                                "type": "percent",
                                "value": parsed["value"],
                                "new_payment": new_payment,
                                "comment": parsed["original"]
                            }
                            if DEBUG_MODE: logger.debug(f"ðŸ“ Applied manager comment for {key}: {parsed['value']}% = {new_payment:.2f}â‚½")
                        elif parsed["type"] == "fixed":
                            # Set fixed payment
                            record["service_payment"] = parsed["value"]
                            record["manager_override"] = True
                            record["manager_override_type"] = "fixed"
                            record["manager_override_value"] = parsed["value"]
                            applied_manager_comments[key] = {
                                "type": "fixed",
                                "value": parsed["value"],
                                "new_payment": parsed["value"],
                                "comment": parsed["original"]
                            }
                            if DEBUG_MODE: logger.debug(f"ðŸ“ Applied manager comment for {key}: fixed {parsed['value']}â‚½")
        
        # Now proceed with calculation (similar to /calculate endpoint)
        # Use default config and calculate
        config = DEFAULT_CONFIG.copy()
        
        # Add Yandex Fuel data to config
        yandex_fuel = session.get("yandex_fuel", {})
        config["yandex_fuel"] = yandex_fuel
        logger.info("ðŸ’¾ /api/apply-review: yandex_fuel from session: {list(yandex_fuel.keys()) if yandex_fuel else 'EMPTY'}")
        
        name_map = session.get("name_map", {})
        
        calculated_data = []
        for row in modified_records:
            calc_row = await calculate_row(row, config, {})

            # For reverted records, restore the old calculation values (including manual edits)
            # This implements "Ð’Ð°Ñ€Ð¸Ð°Ð½Ñ‚ B" - keeping old version values
            if row.get("is_reverted") and "_old_calc_total" in row:
                calc_row["total"] = row["_old_calc_total"]
                calc_row["fuel_payment"] = row.get("_old_calc_fuel", 0)
                calc_row["transport"] = row.get("_old_calc_transport", 0)
                logger.info("âœ… Restored old calc values for {row.get('order', '')[:30]}: total={calc_row['total']}")

            # For restored records (deleted items brought back), preserve their saved total/fuel/transport
            # This is critical for extra rows like "ÐŸÐµÑ€ÐµÐ¿Ð»Ð°Ñ‚Ð°" which have no service_payment
            if row.get("is_restored"):
                if row.get("total", 0) != 0:
                    calc_row["total"] = row["total"]
                if row.get("fuel_payment", 0) != 0:
                    calc_row["fuel_payment"] = row["fuel_payment"]
                if row.get("transport", 0) != 0:
                    calc_row["transport"] = row["transport"]
                logger.info("âœ… Preserved restored row values for {row.get('order', '')[:40]}: total={calc_row['total']}")

            calculated_data.append(calc_row)
        
        calculated_data.sort(key=lambda x: normalize_worker_name(x.get("worker", ""), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""))
        
        # Save to database
        period = session["period"]
        period_id = await get_or_create_period(period)
        upload_id = await create_upload(period_id, config)
        
        # First pass: save all orders and collect totals
        worker_totals = {}
        
        for row in calculated_data:
            worker = row.get("worker", "")
            if not worker or pd.isna(worker):
                continue
            
            is_worker_total = row.get("is_worker_total", False)
            
            # Skip worker total rows - we'll calculate totals from orders
            if is_worker_total:
                continue
            
            # Save individual order
            order_text = str(row.get("order", ""))

            # Use order_code from record if exists (for restored extra rows)
            # Otherwise extract from order text
            order_code = row.get("order_code", "")
            if not order_code:
                match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                if match:
                    order_code = match.group(0)


            # Extract order date from text (format: "ÐšÐÐ£Ð¢-001904, 21.12.2025, ...")
            order_date = None
            date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', order_text)
            if date_match:
                try:
                    day, month, year = date_match.groups()
                    order_date = datetime(int(year), int(month), int(day))
                except ValueError:
                    pass  # Invalid date
            # Use address from record if exists (for restored rows)
            # Otherwise extract from order text
            address = row.get("address", "")
            if not address and ", " in order_text:
                parts = order_text.split(", ", 1)
                if len(parts) > 1:
                    address = parts[1].split("\n")[0][:100]
            
            base_worker = worker.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
            # Check is_client_payment from record first, then from worker name
            is_client = row.get("is_client_payment", False) or "(Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)" in worker
            is_extra = row.get("is_extra_row", False)
            
            order_data = {
                "worker": base_worker,
                "order_code": order_code,
                "order": order_text[:500],
                "order_date": order_date,
                "address": address,
                "is_client_payment": is_client,
                "is_extra_row": is_extra,
                "revenue_total": float(row.get("revenue_total", 0) or 0),
                "revenue_services": float(row.get("revenue_services", 0) or 0),
                "diagnostic": float(row.get("diagnostic", 0) or 0),
                "diagnostic_payment": float(row.get("diagnostic_payment", 0) or 0),
                "specialist_fee": float(row.get("specialist_fee", 0) or 0),
                "additional_expenses": float(row.get("additional_expenses", 0) or 0),
                "service_payment": float(row.get("service_payment", 0) or 0),
                "percent": parse_percent(row.get("percent", 0)),
                "manager_comment": row.get("manager_comment", None)
            }
            order_id = await save_order(upload_id, order_data)
            
            # Save calculation for this order
            total_val = float(row.get("total", 0) or 0)
            calc_data = {
                "worker": base_worker,
                "fuel_payment": float(row.get("fuel_payment", 0) or 0),
                "transport": float(row.get("transport", 0) or 0),
                "diagnostic_50": float(row.get("diagnostic_50", 0) or 0),
                "total": total_val
            }
            await save_calculation(upload_id, order_id, calc_data)
            
            # Accumulate totals per worker
            if base_worker not in worker_totals:
                worker_totals[base_worker] = {"company": 0, "client": 0, "company_count": 0, "client_count": 0}

            if is_client:
                worker_totals[base_worker]["client"] += total_val
                worker_totals[base_worker]["client_count"] += 1
            else:
                worker_totals[base_worker]["company"] += total_val
                worker_totals[base_worker]["company_count"] += 1

        # Save worker totals
        for worker, totals in worker_totals.items():
            await save_worker_total(
                upload_id=upload_id,
                worker=worker,
                total=totals["company"] + totals["client"],
                orders_count=totals["company_count"] + totals["client_count"],
                fuel=0,
                transport=0,
                company_amount=totals["company"],
                client_amount=totals["client"],
                company_orders_count=totals["company_count"],
                client_orders_count=totals["client_count"]
            )
        
        # Save Yandex fuel deductions as manual edits (for history tracking)
        yandex_fuel = config.get("yandex_fuel", {})
        if yandex_fuel:
            from database import save_manual_edit
            for worker, deduction in yandex_fuel.items():
                if deduction and deduction > 0:
                    # Get period name for the order_code field
                    period_name = session.get("period", "")
                    # Determine month from period (e.g., "01-15.12.25" -> "Ð”ÐµÐºÐ°Ð±Ñ€ÑŒ")
                    month_names = {
                        "01": "Ð¯Ð½Ð²Ð°Ñ€ÑŒ", "02": "Ð¤ÐµÐ²Ñ€Ð°Ð»ÑŒ", "03": "ÐœÐ°Ñ€Ñ‚", "04": "ÐÐ¿Ñ€ÐµÐ»ÑŒ",
                        "05": "ÐœÐ°Ð¹", "06": "Ð˜ÑŽÐ½ÑŒ", "07": "Ð˜ÑŽÐ»ÑŒ", "08": "ÐÐ²Ð³ÑƒÑÑ‚",
                        "09": "Ð¡ÐµÐ½Ñ‚ÑÐ±Ñ€ÑŒ", "10": "ÐžÐºÑ‚ÑÐ±Ñ€ÑŒ", "11": "ÐÐ¾ÑÐ±Ñ€ÑŒ", "12": "Ð”ÐµÐºÐ°Ð±Ñ€ÑŒ"
                    }
                    month_num = period_name.split(".")[-2] if "." in period_name else ""
                    month_name = month_names.get(month_num, "")
                    
                    await save_manual_edit(
                        upload_id=upload_id,
                        order_id=None,
                        calculation_id=None,
                        order_code=f"Ð’Ñ‹Ñ‡ÐµÑ‚ Ð¯Ð½Ð´ÐµÐºÑ Ð·Ð°Ð¿Ñ€Ð°Ð²ÐºÐ¸ ({month_name})",
                        worker=worker,
                        address="",
                        field_name="YANDEX_FUEL",
                        old_value=deduction,
                        new_value=-deduction,
                        period_status="DRAFT"
                    )
                    if DEBUG_MODE: logger.debug("â›½ Saved Yandex fuel deduction for {worker}: -{deduction}â‚½")
        
        # Compare with previous upload and save changes
        prev_upload_id = await get_previous_upload(period_id, upload_id)
        if prev_upload_id:
            changes_dict = await compare_uploads(prev_upload_id, upload_id)
            # Process added orders
            for change in changes_dict.get("added", []):
                await save_change(upload_id, change.get("order_code"), change.get("worker"), "added")
            # Process deleted orders
            for change in changes_dict.get("deleted", []):
                await save_change(upload_id, change.get("order_code"), change.get("worker"), "deleted")
            # Process modified orders
            for change in changes_dict.get("modified", []):
                for field_change in change.get("changes", []):
                    await save_change(
                        upload_id, change.get("order_code"), change.get("worker"), 
                        "modified", field_change.get("field"),
                        str(field_change.get("old", "")), str(field_change.get("new", ""))
                    )
        
        # Cleanup session
        del session_data[session_id]
        
        return JSONResponse({
            "success": True,
            "period_id": period_id,
            "upload_id": upload_id
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.post("/api/process-first-upload")
async def process_first_upload(request: Request):
    """Process first upload for a period (no previous version to compare)"""
    try:
        data = await request.json()
        session_id = data.get("session_id")
        
        if session_id not in session_data:
            return JSONResponse({"success": False, "error": "Ð¡ÐµÑÑÐ¸Ñ Ð¸ÑÑ‚ÐµÐºÐ»Ð°"})
        
        session = session_data[session_id]
        combined_records = session.get("combined", [])
        
        # Use default config and add yandex_fuel
        config = DEFAULT_CONFIG.copy()
        yandex_fuel = session.get("yandex_fuel", {})
        config["yandex_fuel"] = yandex_fuel
        logger.info("ðŸ’¾ /api/process-first-upload: yandex_fuel from session: {list(yandex_fuel.keys()) if yandex_fuel else 'EMPTY'}")
        
        name_map = session.get("name_map", {})
        
        calculated_data = []
        for row in combined_records:
            calc_row = await calculate_row(row, config, {})
            calculated_data.append(calc_row)
        
        calculated_data.sort(key=lambda x: normalize_worker_name(x.get("worker", ""), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""))
        
        # Save to database
        period = session["period"]
        period_id = await get_or_create_period(period)
        upload_id = await create_upload(period_id, config)
        
        # First pass: save all orders and collect totals
        worker_totals = {}
        
        for row in calculated_data:
            worker = row.get("worker", "")
            if not worker or pd.isna(worker):
                continue
            
            is_worker_total = row.get("is_worker_total", False)
            
            # Skip worker total rows - we'll calculate totals from orders
            if is_worker_total:
                continue
            
            # Save individual order
            order_text = str(row.get("order", ""))

            # Use order_code from record if exists (for restored extra rows)
            # Otherwise extract from order text
            order_code = row.get("order_code", "")
            if not order_code:
                match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                if match:
                    order_code = match.group(0)


            # Extract order date from text (format: "ÐšÐÐ£Ð¢-001904, 21.12.2025, ...")
            order_date = None
            date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', order_text)
            if date_match:
                try:
                    day, month, year = date_match.groups()
                    order_date = datetime(int(year), int(month), int(day))
                except ValueError:
                    pass  # Invalid date
            # Use address from record if exists (for restored rows)
            # Otherwise extract from order text
            address = row.get("address", "")
            if not address and ", " in order_text:
                parts = order_text.split(", ", 1)
                if len(parts) > 1:
                    address = parts[1].split("\n")[0][:100]
            
            base_worker = worker.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
            # Check is_client_payment from record first, then from worker name
            is_client = row.get("is_client_payment", False) or "(Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)" in worker
            is_extra = row.get("is_extra_row", False)
            
            order_data = {
                "worker": base_worker,
                "order_code": order_code,
                "order": order_text[:500],
                "order_date": order_date,
                "address": address,
                "is_client_payment": is_client,
                "is_extra_row": is_extra,
                "revenue_total": float(row.get("revenue_total", 0) or 0),
                "revenue_services": float(row.get("revenue_services", 0) or 0),
                "diagnostic": float(row.get("diagnostic", 0) or 0),
                "diagnostic_payment": float(row.get("diagnostic_payment", 0) or 0),
                "specialist_fee": float(row.get("specialist_fee", 0) or 0),
                "additional_expenses": float(row.get("additional_expenses", 0) or 0),
                "service_payment": float(row.get("service_payment", 0) or 0),
                "percent": parse_percent(row.get("percent", 0)),
                "manager_comment": row.get("manager_comment", None)
            }
            order_id = await save_order(upload_id, order_data)
            
            # Save calculation for this order
            total_val = float(row.get("total", 0) or 0)
            calc_data = {
                "worker": base_worker,
                "fuel_payment": float(row.get("fuel_payment", 0) or 0),
                "transport": float(row.get("transport", 0) or 0),
                "diagnostic_50": float(row.get("diagnostic_50", 0) or 0),
                "total": total_val
            }
            await save_calculation(upload_id, order_id, calc_data)
            
            # Accumulate totals per worker
            if base_worker not in worker_totals:
                worker_totals[base_worker] = {"company": 0, "client": 0, "company_count": 0, "client_count": 0}

            if is_client:
                worker_totals[base_worker]["client"] += total_val
                worker_totals[base_worker]["client_count"] += 1
            else:
                worker_totals[base_worker]["company"] += total_val
                worker_totals[base_worker]["company_count"] += 1

        # Save worker totals
        for worker, totals in worker_totals.items():
            await save_worker_total(
                upload_id=upload_id,
                worker=worker,
                total=totals["company"] + totals["client"],
                orders_count=totals["company_count"] + totals["client_count"],
                fuel=0,
                transport=0,
                company_amount=totals["company"],
                client_amount=totals["client"],
                company_orders_count=totals["company_count"],
                client_orders_count=totals["client_count"]
            )
        
        # Save Yandex fuel deductions as manual edits (for history tracking)
        yandex_fuel = config.get("yandex_fuel", {})
        if yandex_fuel:
            from database import save_manual_edit
            for worker, deduction in yandex_fuel.items():
                if deduction and deduction > 0:
                    # Get period name for the order_code field
                    period_name = session.get("period", "")
                    # Determine month from period (e.g., "01-15.12.25" -> "Ð”ÐµÐºÐ°Ð±Ñ€ÑŒ")
                    month_names = {
                        "01": "Ð¯Ð½Ð²Ð°Ñ€ÑŒ", "02": "Ð¤ÐµÐ²Ñ€Ð°Ð»ÑŒ", "03": "ÐœÐ°Ñ€Ñ‚", "04": "ÐÐ¿Ñ€ÐµÐ»ÑŒ",
                        "05": "ÐœÐ°Ð¹", "06": "Ð˜ÑŽÐ½ÑŒ", "07": "Ð˜ÑŽÐ»ÑŒ", "08": "ÐÐ²Ð³ÑƒÑÑ‚",
                        "09": "Ð¡ÐµÐ½Ñ‚ÑÐ±Ñ€ÑŒ", "10": "ÐžÐºÑ‚ÑÐ±Ñ€ÑŒ", "11": "ÐÐ¾ÑÐ±Ñ€ÑŒ", "12": "Ð”ÐµÐºÐ°Ð±Ñ€ÑŒ"
                    }
                    month_num = period_name.split(".")[-2] if "." in period_name else ""
                    month_name = month_names.get(month_num, "")
                    
                    await save_manual_edit(
                        upload_id=upload_id,
                        order_id=None,
                        calculation_id=None,
                        order_code=f"Ð’Ñ‹Ñ‡ÐµÑ‚ Ð¯Ð½Ð´ÐµÐºÑ Ð·Ð°Ð¿Ñ€Ð°Ð²ÐºÐ¸ ({month_name})",
                        worker=worker,
                        address="",
                        field_name="YANDEX_FUEL",
                        old_value=deduction,
                        new_value=-deduction,
                        period_status="DRAFT"
                    )
                    if DEBUG_MODE: logger.debug("â›½ Saved Yandex fuel deduction for {worker}: -{deduction}â‚½")
        
        # Cleanup session
        del session_data[session_id]
        
        return JSONResponse({
            "success": True,
            "period_id": period_id,
            "upload_id": upload_id
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.post("/preview")
async def preview_calculation(
    session_id: str = Form(...),
    config_json: str = Form(...),
    days_json: str = Form(...),
    extra_rows_json: str = Form(...)
):
    """Preview calculation without saving to database"""
    try:
        if session_id not in session_data:
            raise HTTPException(status_code=400, detail="Session expired")
        
        session = session_data[session_id]
        config = json.loads(config_json)
        days_map = json.loads(days_json)
        extra_rows = json.loads(extra_rows_json)
        
        full_config = {**DEFAULT_CONFIG, **config}
        name_map = session.get("name_map", {})
        
        # Add Yandex Fuel data to config for report generation
        yandex_fuel = session.get("yandex_fuel", {})
        full_config["yandex_fuel"] = yandex_fuel
        
        calculated_data = []
        for row in session["combined"]:
            calc_row = await calculate_row(row, full_config, days_map)
            calculated_data.append(calc_row)
        
        for worker, rows in extra_rows.items():
            for extra in rows:
                calculated_data.append({
                    "worker": normalize_worker_name(worker, name_map),
                    "order": extra.get("description", ""),
                    "order_id": f"extra_{worker}_{extra.get('description', '')[:20]}",
                    "revenue_total": "",
                    "revenue_services": "",
                    "diagnostic": "",
                    "diagnostic_payment": "",
                    "specialist_fee": "",
                    "additional_expenses": "",
                    "service_payment": "",
                    "percent": "",
                    "is_over_10k": False,
                    "is_client_payment": False,
                    "is_worker_total": False,
                    "is_extra_row": True,
                    "fuel_payment": "",
                    "transport": "",
                    "diagnostic_50": "",
                    "total": float(extra.get("amount", 0))
                })
        
        calculated_data.sort(key=lambda x: normalize_worker_name(x.get("worker", ""), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""))
        
        # Generate unique IDs for each row for deletion
        preview_rows = []
        for idx, row in enumerate(calculated_data):
            order_code = ""
            order_text = row.get("order", "")
            match = re.search(r'((?:ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+)', order_text)
            if match:
                order_code = match.group(1)
            
            # Extract address
            address = extract_address_from_order(order_text)
            if not address:
                address = order_text[:50] + "..." if len(order_text) > 50 else order_text
            
            preview_rows.append({
                "id": idx,
                "worker": row.get("worker", "").replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""),
                "order_code": order_code,
                "address": address,
                "revenue_total": row.get("revenue_total", ""),
                "revenue_services": row.get("revenue_services", ""),
                "service_payment": row.get("service_payment", ""),
                "percent": row.get("percent", ""),
                "fuel_payment": row.get("fuel_payment", ""),
                "transport": row.get("transport", ""),
                "total": row.get("total", 0),
                "is_client_payment": row.get("is_client_payment", False),
                "is_over_10k": row.get("is_over_10k", False),
                "is_extra_row": row.get("is_extra_row", False)
            })
        
        # Store calculated data in session for later finalization
        session["calculated_data"] = calculated_data
        session["config"] = full_config
        
        # Group by worker for summary
        workers_summary = {}
        for row in preview_rows:
            worker = row["worker"]
            if worker not in workers_summary:
                workers_summary[worker] = {
                    "total": 0,
                    "count": 0,
                    "fuel": 0,
                    "transport": 0
                }
            total = row.get("total", 0)
            if isinstance(total, (int, float)):
                workers_summary[worker]["total"] += total
            workers_summary[worker]["count"] += 1
            fuel = row.get("fuel_payment", 0)
            if isinstance(fuel, (int, float)):
                workers_summary[worker]["fuel"] += fuel
            transport = row.get("transport", 0)
            if isinstance(transport, (int, float)):
                workers_summary[worker]["transport"] += transport
        
        alarms = generate_alarms(calculated_data, full_config)
        
        return JSONResponse({
            "success": True,
            "rows": preview_rows,
            "workers_summary": workers_summary,
            "alarms": alarms,
            "period": session["period"]
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/calculate")
async def calculate_salaries(
    session_id: str = Form(...),
    config_json: str = Form(...),
    days_json: str = Form(...),
    extra_rows_json: str = Form(...),
    deleted_rows_json: str = Form("[]")
):
    """Calculate all salaries and generate files"""
    try:
        if session_id not in session_data:
            raise HTTPException(status_code=400, detail="Session expired")
        
        session = session_data[session_id]
        deleted_rows_raw = json.loads(deleted_rows_json)
        deleted_rows = set(int(x) for x in deleted_rows_raw)
        
        if deleted_rows:
            logger.info(f"ðŸ—‘ï¸ Ð£Ð´Ð°Ð»ÑÐµÐ¼ ÑÑ‚Ñ€Ð¾ÐºÐ¸ Ñ ID: {sorted(deleted_rows)}")
        
        # Use pre-calculated data from preview if available
        if "calculated_data" in session:
            all_calculated = session["calculated_data"]
            full_config = session.get("config", DEFAULT_CONFIG)
            
            # Ensure yandex_fuel is in config (might be missing if preview was skipped)
            if "yandex_fuel" not in full_config:
                full_config = {**full_config, "yandex_fuel": session.get("yandex_fuel", {})}
            
            # Filter out deleted rows
            calculated_data = []
            for idx, row in enumerate(all_calculated):
                if idx in deleted_rows:
                    order_info = row.get("order", "")[:50]
                    worker = row.get("worker", "")
                    if DEBUG_MODE: logger.debug(f"  âŒ Ð£Ð´Ð°Ð»ÑÐµÐ¼: {worker} - {order_info}")
                    continue
                calculated_data.append(row)
        else:
            # Fallback: recalculate (shouldn't normally happen)
            config = json.loads(config_json)
            days_map = json.loads(days_json)
            extra_rows = json.loads(extra_rows_json)
            
            full_config = {**DEFAULT_CONFIG, **config}
            name_map = session.get("name_map", {})
            
            # Add Yandex Fuel data to config
            yandex_fuel = session.get("yandex_fuel", {})
            full_config["yandex_fuel"] = yandex_fuel
            
            calculated_data = []
            for idx, row in enumerate(session["combined"]):
                calc_row = await calculate_row(row, full_config, days_map)
                calculated_data.append(calc_row)
            
            for worker, rows in extra_rows.items():
                for extra in rows:
                    calculated_data.append({
                        "worker": normalize_worker_name(worker, name_map),
                        "order": extra.get("description", ""),
                        "revenue_total": "",
                        "revenue_services": "",
                        "diagnostic": "",
                        "diagnostic_payment": "",
                        "specialist_fee": "",
                        "additional_expenses": "",
                        "service_payment": "",
                        "percent": "",
                        "is_over_10k": False,
                        "is_client_payment": False,
                        "is_worker_total": False,
                        "is_extra_row": True,
                        "fuel_payment": "",
                        "transport": "",
                        "diagnostic_50": "",
                        "total": float(extra.get("amount", 0))
                    })
            
            # Sort same as preview
            name_map = session.get("name_map", {})
            calculated_data.sort(key=lambda x: normalize_worker_name(x.get("worker", ""), name_map).replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""))
            
            # Now filter deleted
            calculated_data = [row for idx, row in enumerate(calculated_data) if idx not in deleted_rows]
        
        alarms = generate_alarms(calculated_data, full_config)
        
        period = session["period"]
        workers = session["workers"]
        
        # Archive 1: Full reports (for accounting)
        zip_full = BytesIO()
        with zipfile.ZipFile(zip_full, "w", zipfile.ZIP_DEFLATED) as zf:
            main_report = create_excel_report(calculated_data, period, full_config, for_workers=False)
            zf.writestr(f"ÐžÐ±Ñ‰Ð¸Ð¹_Ð¾Ñ‚Ñ‡ÐµÑ‚ {period}.xlsx", main_report)
            
            for worker in workers:
                worker_surname = worker.split()[0] if worker else "Unknown"
                worker_report = create_worker_report(calculated_data, worker, period, full_config, for_workers=False)
                zf.writestr(f"{worker_surname} {period}.xlsx", worker_report)
        
        zip_full.seek(0)
        
        # Archive 2: Simplified reports (for workers - hidden columns)
        zip_workers = BytesIO()
        with zipfile.ZipFile(zip_workers, "w", zipfile.ZIP_DEFLATED) as zf:
            main_report = create_excel_report(calculated_data, period, full_config, for_workers=True)
            zf.writestr(f"ÐžÐ±Ñ‰Ð¸Ð¹_Ð¾Ñ‚Ñ‡ÐµÑ‚ {period}.xlsx", main_report)
            
            for worker in workers:
                worker_surname = worker.split()[0] if worker else "Unknown"
                worker_report = create_worker_report(calculated_data, worker, period, full_config, for_workers=True)
                zf.writestr(f"{worker_surname} {period}.xlsx", worker_report)
        
        zip_workers.seek(0)
        
        # Save both archives
        temp_path_full = f"/tmp/salary_report_{session_id}_full.zip"
        temp_path_workers = f"/tmp/salary_report_{session_id}_workers.zip"
        
        with open(temp_path_full, "wb") as f:
            f.write(zip_full.getvalue())
        
        with open(temp_path_workers, "wb") as f:
            f.write(zip_workers.getvalue())
        
        session_data[session_id]["alarms"] = alarms
        
        # ===== SAVE TO DATABASE =====
        try:
            if database:
                # Debug: check yandex_fuel before saving
                yf = full_config.get("yandex_fuel", {})
                logger.info("ðŸ’¾ /calculate: yandex_fuel in full_config: {list(yf.keys()) if yf else 'EMPTY'}")
                
                # 1. Get or create period
                period_id = await get_or_create_period(period)
                
                # 2. Create upload
                upload_id = await create_upload(period_id, full_config)
                
                # 3. Check for previous upload and compare
                prev_upload_id = await get_previous_upload(period_id, 
                    (await get_period_details(period_id))["uploads"][0]["version"] if (await get_period_details(period_id))["uploads"] else 1
                )
                
                # 4. Save orders and calculations - ONLY for valid workers
                order_id_map = {}  # To map order to its DB id
                for row in calculated_data:
                    # Skip non-worker groups (Ð”Ð¾ÑÑ‚Ð°Ð²ÐºÐ°, ÐŸÐ¾Ð¼Ð¾Ñ‰Ð½Ð¸Ðº, etc.)
                    worker = normalize_worker_name(row.get("worker", "").replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""))
                    if not is_valid_worker_name(worker):
                        continue
                    
                    is_extra = row.get("is_extra_row", False)
                    
                    # Extract order code from order text (for regular rows)
                    order_text = row.get("order", "")
                    order_code_match = re.search(r'(ÐšÐÐ£Ð¢|Ð˜Ð‘Ð£Ð¢|Ð¢Ð”Ð£Ð¢)-\d+', order_text)
                    order_code = order_code_match.group(0) if order_code_match else ""
                    
                    # For extra rows, use description as order text
                    if is_extra:
                        order_code = "Ð”ÐžÐŸÐ›ÐÐ¢Ð"  # Special code for extra rows
                    
                    # Save order
                    order_data = {
                        "worker": row.get("worker", ""),
                        "order_code": order_code,
                        "order": order_text,
                        "address": extract_address_from_order(order_text) if not is_extra else order_text,
                        "revenue_total": row.get("revenue_total", 0) if not is_extra else 0,
                        "revenue_services": row.get("revenue_services", 0) if not is_extra else 0,
                        "diagnostic": row.get("diagnostic", 0) if not is_extra else 0,
                        "diagnostic_payment": row.get("diagnostic_payment", 0) if not is_extra else 0,
                        "specialist_fee": row.get("specialist_fee", 0) if not is_extra else 0,
                        "additional_expenses": row.get("additional_expenses", 0) if not is_extra else 0,
                        "service_payment": row.get("service_payment", 0) if not is_extra else 0,
                        "percent": row.get("percent", "") if not is_extra else "",
                        "is_client_payment": row.get("is_client_payment", False),
                        "is_over_10k": row.get("is_over_10k", False),
                        "is_extra_row": is_extra,
                    }
                    order_id = await save_order(upload_id, order_data)
                    order_id_map[order_text] = order_id
                    
                    # Save calculation
                    calc_data = {
                        "worker": row.get("worker", ""),
                        "fuel_payment": row.get("fuel_payment", 0) if not is_extra else 0,
                        "transport": row.get("transport", 0) if not is_extra else 0,
                        "diagnostic_50": row.get("diagnostic_50", 0) if not is_extra else 0,
                        "total": row.get("total", 0),
                    }
                    await save_calculation(upload_id, order_id, calc_data)
                
                # 5. Calculate and save worker totals - ONLY for valid workers
                worker_totals_dict = {}
                for row in calculated_data:
                    worker = normalize_worker_name(row.get("worker", "").replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", ""))
                    
                    # Skip non-worker groups (Ð”Ð¾ÑÑ‚Ð°Ð²ÐºÐ°, ÐŸÐ¾Ð¼Ð¾Ñ‰Ð½Ð¸Ðº, etc.)
                    if not is_valid_worker_name(worker):
                        continue
                    
                    if worker not in worker_totals_dict:
                        worker_totals_dict[worker] = {
                            "total": 0,
                            "company_total": 0,
                            "client_total": 0,
                            "count": 0,
                            "company_count": 0,
                            "client_count": 0,
                            "fuel": 0,
                            "transport": 0
                        }
                    
                    total = row.get("total", 0)
                    is_client = row.get("is_client_payment", False)
                    
                    if isinstance(total, (int, float)):
                        worker_totals_dict[worker]["total"] += total
                        if is_client:
                            worker_totals_dict[worker]["client_total"] += total
                            worker_totals_dict[worker]["client_count"] += 1
                        else:
                            worker_totals_dict[worker]["company_total"] += total
                            worker_totals_dict[worker]["company_count"] += 1
                    
                    worker_totals_dict[worker]["count"] += 1
                    
                    fuel = row.get("fuel_payment", 0)
                    if isinstance(fuel, (int, float)):
                        worker_totals_dict[worker]["fuel"] += fuel
                    
                    transport = row.get("transport", 0)
                    if isinstance(transport, (int, float)):
                        worker_totals_dict[worker]["transport"] += transport
                
                for worker, totals in worker_totals_dict.items():
                    await save_worker_total(
                        upload_id, 
                        worker, 
                        totals["total"],
                        totals["count"],
                        totals["fuel"],
                        totals["transport"],
                        totals["company_total"],
                        totals["client_total"],
                        totals["company_count"],
                        totals["client_count"]
                    )
                
                # 6. Compare with previous upload if exists
                if prev_upload_id:
                    changes_dict = await compare_uploads(prev_upload_id, upload_id)
                    # Process added orders
                    for change in changes_dict.get("added", []):
                        await save_change(upload_id, change.get("order_code"), change.get("worker"), "added")
                    # Process deleted orders
                    for change in changes_dict.get("deleted", []):
                        await save_change(upload_id, change.get("order_code"), change.get("worker"), "deleted")
                    # Process modified orders
                    for change in changes_dict.get("modified", []):
                        for field_change in change.get("changes", []):
                            await save_change(
                                upload_id, 
                                change.get("order_code"), 
                                change.get("worker"), 
                                "modified",
                                field_change.get("field"),
                                str(field_change.get("old", "")),
                                str(field_change.get("new", ""))
                            )
                
                logger.info("âœ… Saved to database: period={period}, upload_id={upload_id}")
        except Exception as db_error:
            logger.warning("âš ï¸ Database save error (non-critical): {db_error}")
            # Don't fail the request if DB save fails
        
        return JSONResponse({
            "success": True,
            "download_url_full": f"/download/{session_id}/full",
            "download_url_workers": f"/download/{session_id}/workers",
            "alarms": alarms
        })
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/download/{session_id}/{archive_type}")
async def download_report(session_id: str, archive_type: str):
    """Download generated ZIP file
    
    Args:
        archive_type: 'full' for accounting reports, 'workers' for simplified reports
    """
    if archive_type not in ["full", "workers"]:
        raise HTTPException(status_code=400, detail="Invalid archive type")
    
    temp_path = f"/tmp/salary_report_{session_id}_{archive_type}.zip"
    if not os.path.exists(temp_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    period = session_data.get(session_id, {}).get("period", "report")
    
    if archive_type == "full":
        filename = f"Ð—Ð°Ñ€Ð¿Ð»Ð°Ñ‚Ð°_{period}.zip"
    else:
        filename = f"Ð”Ð»Ñ_Ð¼Ð¾Ð½Ñ‚Ð°Ð¶Ð½Ð¸ÐºÐ¾Ð²_{period}.zip"
    
    return FileResponse(
        temp_path,
        media_type="application/zip",
        filename=filename
    )


# ============== HISTORY API ENDPOINTS ==============

@app.get("/history")
async def history_page(request: Request):
    """History page - view all periods by month"""
    user = get_current_user(request)
    response = templates.TemplateResponse("history.html", {"request": request, "user": user})
    # Prevent browser caching so back button shows fresh data
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/api/periods")
async def api_get_periods():
    """Get all periods grouped by month"""
    try:
        periods = await get_all_periods()
        
        # Enrich periods with total_amount and latest_upload_id
        enriched_periods = []
        for p in periods:
            period_details = await get_period_details(p["id"])
            total_amount = 0
            company_amount = 0
            client_amount = 0
            latest_upload_id = None
            
            if period_details and period_details.get("uploads"):
                latest_upload = period_details["uploads"][0]
                latest_upload_id = latest_upload["id"]
                
                upload_details = await get_upload_details(latest_upload_id)
                if upload_details:
                    worker_totals = upload_details.get("worker_totals", [])
                    for wt in worker_totals:
                        company_amount += wt.get("company_amount", 0) or 0
                        client_amount += wt.get("client_amount", 0) or 0
                    total_amount = company_amount + client_amount
            
            # Convert datetime fields to strings for JSON serialization
            created_at = p.get("created_at")
            sent_at = p.get("sent_at")
            paid_at = p.get("paid_at")
            
            enriched_periods.append({
                "id": p.get("id"),
                "name": p.get("name"),
                "month": p.get("month"),
                "year": p.get("year"),
                "status": p.get("status", "draft"),
                "created_at": str(created_at) if created_at else None,
                "sent_at": str(sent_at) if sent_at else None,
                "paid_at": str(paid_at) if paid_at else None,
                "total_amount": total_amount,
                "company_amount": company_amount,
                "client_amount": client_amount,
                "latest_upload_id": latest_upload_id
            })
        
        # Group by month
        months = {}
        for p in enriched_periods:
            month_key = p["month"]
            if month_key not in months:
                months[month_key] = {
                    "month": month_key,
                    "year": p["year"],
                    "periods": []
                }
            months[month_key]["periods"].append(p)
        
        return JSONResponse({
            "success": True,
            "months": list(months.values())
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/period/{period_id}")
async def api_get_period(period_id: int):
    """Get period details with uploads"""
    try:
        details = await get_period_details(period_id)
        if not details:
            raise HTTPException(status_code=404, detail="Period not found")
        
        # Convert datetime fields to strings
        def serialize_datetime(obj):
            if isinstance(obj, dict):
                return {k: serialize_datetime(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [serialize_datetime(item) for item in obj]
            elif hasattr(obj, 'isoformat'):  # datetime object
                return obj.isoformat()
            return obj
        
        serialized = serialize_datetime(details)
        
        return JSONResponse({
            "success": True,
            "data": serialized
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/upload/{upload_id}")
async def api_get_upload(upload_id: int):
    """Get upload details with worker totals"""
    try:
        details = await get_upload_details(upload_id)
        if not details:
            raise HTTPException(status_code=404, detail="Upload not found")
        
        # Convert datetime fields to strings
        def serialize_datetime(obj):
            if isinstance(obj, dict):
                return {k: serialize_datetime(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [serialize_datetime(item) for item in obj]
            elif hasattr(obj, 'isoformat'):  # datetime object
                return obj.isoformat()
            return obj
        
        serialized = serialize_datetime(details)
        
        return JSONResponse({
            "success": True,
            "data": serialized
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/upload/{upload_id}/worker/{worker}")
async def api_get_worker_orders(upload_id: int, worker: str):
    """Get all orders for a worker"""
    try:
        from urllib.parse import unquote
        worker_decoded = unquote(worker)
        
        # Get orders for this worker (both regular and client payment)
        orders = await get_worker_orders(upload_id, worker_decoded)
        orders_client = await get_worker_orders(upload_id, f"{worker_decoded} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)")
        
        # Get worker totals
        upload_details = await get_upload_details(upload_id)
        worker_totals = upload_details.get("worker_totals", []) if upload_details else []
        
        worker_total = None
        for wt in worker_totals:
            if wt.get("worker") == worker_decoded:
                worker_total = wt
                break
        
        return JSONResponse({
            "success": True,
            "data": {
                "worker": worker_decoded,
                "orders": orders + orders_client,
                "worker_total": worker_total
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/worker-report/{upload_id}/{worker}")
async def api_worker_report(upload_id: int, worker: str):
    """Download Excel report for a specific worker"""
    try:
        from urllib.parse import unquote
        worker_decoded = unquote(worker)
        
        # Get upload details
        upload_details = await get_upload_details(upload_id)
        if not upload_details:
            raise HTTPException(status_code=404, detail="Upload not found")
        
        # Get period info - upload_details contains period_id directly
        period_id = upload_details.get("period_id")
        
        period_details = await get_period_details(period_id)
        period_name = period_details.get("name", "") if period_details else ""
        
        # Load config from DB (includes yandex_fuel if saved)
        saved_config = upload_details.get("config_json", {}) or {}
        if isinstance(saved_config, str):
            saved_config = json.loads(saved_config)
        report_config = {**DEFAULT_CONFIG, **saved_config}
        
        # Get ALL orders for this upload (needed for proper report generation)
        worker_totals_list = upload_details.get("worker_totals", [])
        
        all_orders = []
        for wt in worker_totals_list:
            worker_orders = await get_worker_orders(upload_id, wt["worker"])
            all_orders.extend(worker_orders)
            worker_orders_client = await get_worker_orders(upload_id, f"{wt['worker']} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)")
            all_orders.extend(worker_orders_client)
        
        # Build calculated_data structure (same as archive generation)
        calculated_data = []
        for order in all_orders:
            # Values from calculations table (already in order from JOIN)
            row = {
                "worker": order["worker"],
                "order": order.get("order_full", "") or order.get("address", ""),
                "order_code": order.get("order_code", ""),
                "address": order.get("address", ""),
                "revenue_total": order.get("revenue_total", 0),
                "revenue_services": order.get("revenue_services", 0),
                "diagnostic": order.get("diagnostic", 0),
                "diagnostic_payment": order.get("diagnostic_payment", 0),
                "specialist_fee": order.get("specialist_fee", 0),
                "additional_expenses": order.get("additional_expenses", 0),
                "service_payment": order.get("service_payment", 0),
                "percent": order.get("percent", ""),
                "is_client_payment": order.get("is_client_payment", False),
                "is_over_10k": order.get("is_over_10k", False),
                "is_extra_row": order.get("is_extra_row", False),
                # Values from calculations table (already in order from JOIN)
                "fuel_payment": order.get("fuel_payment", 0) or 0,
                "transport": order.get("transport", 0) or 0,
                "diagnostic_50": order.get("diagnostic_50", 0) or 0,
                "total": order.get("total", 0) or 0,
            }
            calculated_data.append(row)
        
        # Generate worker report using same function as archive
        report_bytes = create_worker_report(calculated_data, worker_decoded, period_name, report_config, for_workers=True)
        
        # Create safe filename
        worker_surname = worker_decoded.split()[0] if worker_decoded else "worker"
        filename = f"{worker_surname}_{period_name.replace('.', '_')}.xlsx"
        
        return Response(
            content=report_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/months-summary")
async def api_months_summary():
    """Get summary by months for dashboard"""
    try:
        summary = await get_months_summary()
        return JSONResponse({
            "success": True,
            "summary": summary
        })
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/comparison")
async def api_comparison():
    """Get comparison data for all periods and workers"""
    try:
        periods = await get_all_periods()
        
        # Get all unique workers and worker totals for each period
        all_workers = set()
        periods_with_totals = []
        
        for period in periods:
            period_details = await get_period_details(period["id"])
            if period_details and period_details.get("uploads"):
                latest_upload = period_details["uploads"][0]
                upload_details = await get_upload_details(latest_upload["id"])
                
                worker_totals = upload_details.get("worker_totals", [])
                for wt in worker_totals:
                    all_workers.add(wt["worker"])
                
                periods_with_totals.append({
                    "id": period["id"],
                    "name": period["name"],
                    "month": period["month"],
                    "worker_totals": worker_totals
                })
        
        # Group by months
        months_map = {}
        for p in periods_with_totals:
            month = p["month"]
            if month not in months_map:
                months_map[month] = {
                    "month": month,
                    "worker_totals": []
                }
            # Aggregate worker totals by month
            for wt in p["worker_totals"]:
                existing = next((w for w in months_map[month]["worker_totals"] if w["worker"] == wt["worker"]), None)
                if existing:
                    existing["total_amount"] = existing.get("total_amount", 0) + wt.get("total_amount", 0)
                    existing["company_amount"] = existing.get("company_amount", 0) + wt.get("company_amount", 0)
                    existing["client_amount"] = existing.get("client_amount", 0) + wt.get("client_amount", 0)
                else:
                    months_map[month]["worker_totals"].append({
                        "worker": wt["worker"],
                        "total_amount": wt.get("total_amount", 0),
                        "company_amount": wt.get("company_amount", 0),
                        "client_amount": wt.get("client_amount", 0)
                    })
        
        months_list = sorted(months_map.values(), key=lambda x: x["month"], reverse=True)
        
        return JSONResponse({
            "success": True,
            "data": {
                "periods": periods_with_totals,
                "months": months_list,
                "workers": sorted(list(all_workers))
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/comparison/export")
async def api_comparison_export():
    """Export comparison table to Excel"""
    try:
        # Get comparison data
        periods = await get_all_periods()
        
        all_workers = set()
        periods_with_totals = []
        
        for period in periods:
            period_details = await get_period_details(period["id"])
            if period_details and period_details.get("uploads"):
                latest_upload = period_details["uploads"][0]
                upload_details = await get_upload_details(latest_upload["id"])
                
                worker_totals = upload_details.get("worker_totals", [])
                for wt in worker_totals:
                    all_workers.add(wt["worker"])
                
                periods_with_totals.append({
                    "id": period["id"],
                    "name": period["name"],
                    "month": period["month"],
                    "worker_totals": worker_totals
                })
        
        workers = sorted(list(all_workers))
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ð¡Ñ€Ð°Ð²Ð½ÐµÐ½Ð¸Ðµ Ð¿Ð¾ Ð¿ÐµÑ€Ð¸Ð¾Ð´Ð°Ð¼"
        
        # Header style
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        ws.cell(row=1, column=1, value="ÐœÐ¾Ð½Ñ‚Ð°Ð¶Ð½Ð¸Ðº").fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        
        for col, period in enumerate(periods_with_totals, start=2):
            cell = ws.cell(row=1, column=col, value=period["name"])
            cell.fill = header_fill
            cell.font = header_font
        
        total_col = len(periods_with_totals) + 2
        ws.cell(row=1, column=total_col, value="Ð’ÑÐµÐ³Ð¾").fill = header_fill
        ws.cell(row=1, column=total_col).font = header_font
        
        # Data rows
        for row, worker in enumerate(workers, start=2):
            ws.cell(row=row, column=1, value=worker)
            worker_total = 0
            
            for col, period in enumerate(periods_with_totals, start=2):
                wt = next((w for w in period["worker_totals"] if w["worker"] == worker), None)
                value = wt.get("total_amount", 0) if wt else 0
                worker_total += value
                ws.cell(row=row, column=col, value=round(value))
            
            ws.cell(row=row, column=total_col, value=round(worker_total))
        
        # Total row
        total_row = len(workers) + 2
        ws.cell(row=total_row, column=1, value="Ð˜Ð¢ÐžÐ“Ðž").font = Font(bold=True)
        
        for col, period in enumerate(periods_with_totals, start=2):
            period_total = sum(w.get("total_amount", 0) for w in period["worker_totals"])
            ws.cell(row=total_row, column=col, value=round(period_total)).font = Font(bold=True)
        
        grand_total = sum(
            sum(w.get("total_amount", 0) for w in p["worker_totals"])
            for p in periods_with_totals
        )
        ws.cell(row=total_row, column=total_col, value=round(grand_total)).font = Font(bold=True)
        
        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=comparison.xlsx"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/period/{period_id}/download/{archive_type}")
async def download_period_archive(period_id: int, archive_type: str):
    """Download archive for a specific period - generates full archive like step 4"""
    try:
        period_details = await get_period_details(period_id)
        if not period_details or not period_details.get("uploads"):
            raise HTTPException(status_code=404, detail="Period not found")
        
        latest_upload = period_details["uploads"][0]
        upload_details = await get_upload_details(latest_upload["id"])
        
        # Get worker totals
        worker_totals_list = upload_details.get("worker_totals", [])
        workers = [wt["worker"] for wt in worker_totals_list]
        
        # Get all orders for this upload with FRESH calculation data
        all_orders = []
        for wt in worker_totals_list:
            worker_orders = await get_worker_orders(latest_upload["id"], wt["worker"])
            all_orders.extend(worker_orders)
            # Also get client payment orders
            worker_orders_client = await get_worker_orders(latest_upload["id"], f"{wt['worker']} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)")
            all_orders.extend(worker_orders_client)
        
        # Reconstruct data structure from current DB values (calculations joined with orders!)
        calculated_data = []
        for order in all_orders:
            # Note: fuel_payment, transport, etc. are already in order from JOIN
            row = {
                "worker": order["worker"],
                "order": order.get("order_full", "") or order.get("address", ""),
                "order_code": order.get("order_code", ""),
                "address": order.get("address", ""),
                "revenue_total": order.get("revenue_total", 0),
                "revenue_services": order.get("revenue_services", 0),
                "diagnostic": order.get("diagnostic", 0),
                "diagnostic_payment": order.get("diagnostic_payment", 0),
                "specialist_fee": order.get("specialist_fee", 0),
                "additional_expenses": order.get("additional_expenses", 0),
                "service_payment": order.get("service_payment", 0),
                "percent": order.get("percent", ""),
                "is_client_payment": order.get("is_client_payment", False),
                "is_over_10k": order.get("is_over_10k", False),
                "is_extra_row": order.get("is_extra_row", False),
                # Values from calculations table (already in order from JOIN)
                "fuel_payment": order.get("fuel_payment", 0) or 0,
                "transport": order.get("transport", 0) or 0,
                "diagnostic_50": order.get("diagnostic_50", 0) or 0,
                "total": order.get("total", 0) or 0,
            }
            calculated_data.append(row)
        
        if DEBUG_MODE: logger.debug("ðŸ“Š Generating archive from {len(calculated_data)} orders")
        # Debug: show some totals
        for wt in worker_totals_list[:3]:
            worker_data = [r for r in calculated_data if r["worker"].replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "") == wt["worker"]]
            calc_total = sum(r.get("total", 0) for r in worker_data)
            if DEBUG_MODE: logger.debug(f"   {wt['worker']}: {len(worker_data)} orders, calc_total={calc_total}")
        
        period_name = period_details.get("name", f"period_{period_id}")
        for_workers = (archive_type == "workers")
        
        # Load config from DB (includes yandex_fuel if saved)
        saved_config = upload_details.get("config_json", {}) or {}
        if isinstance(saved_config, str):
            saved_config = json.loads(saved_config)
        report_config = {**DEFAULT_CONFIG, **saved_config}
        
        # Debug: check if yandex_fuel is in config
        yandex_fuel = report_config.get("yandex_fuel", {})
        if DEBUG_MODE: logger.debug("ðŸ“Š Download config yandex_fuel: {yandex_fuel}")
        
        # Generate FULL archive with all worker files (like step 4)
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            # Main report
            main_report = create_excel_report(calculated_data, period_name, report_config, for_workers=for_workers)
            main_filename = f"Ð”Ð»Ñ_Ð¼Ð¾Ð½Ñ‚Ð°Ð¶Ð½Ð¸ÐºÐ¾Ð²_{period_name.replace('.', '_')}.xlsx" if for_workers else f"ÐžÐ±Ñ‰Ð¸Ð¹_Ð¾Ñ‚Ñ‡ÐµÑ‚_{period_name.replace('.', '_')}.xlsx"
            zf.writestr(main_filename, main_report)
            
            # Individual worker reports
            for worker in workers:
                worker_surname = worker.split()[0] if worker else "Unknown"
                worker_report = create_worker_report(calculated_data, worker, period_name, report_config, for_workers=for_workers)
                zf.writestr(f"{worker_surname}_{period_name.replace('.', '_')}.xlsx", worker_report)
        
        zip_buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        from urllib.parse import quote
        
        # Use ASCII-safe filename and add UTF-8 encoded filename for modern browsers
        archive_name = f"{'Ð”Ð»Ñ_Ð¼Ð¾Ð½Ñ‚Ð°Ð¶Ð½Ð¸ÐºÐ¾Ð²' if for_workers else 'ÐŸÐ¾Ð»Ð½Ñ‹Ð¹_Ð¾Ñ‚Ñ‡ÐµÑ‚'}_{period_name.replace('.', '_')}.zip"
        ascii_name = f"{'workers' if for_workers else 'full'}_{period_name.replace('.', '_')}.zip"
        
        # RFC 5987 encoding for non-ASCII filenames
        encoded_name = quote(archive_name)
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/comparison")
async def comparison_page(request: Request):
    """Comparison table page"""
    user = get_current_user(request)
    return templates.TemplateResponse("comparison.html", {"request": request, "user": user})


@app.post("/api/order/{order_id}/calculation")
async def create_or_update_order_calculation(order_id: int, request: Request):
    """Create or update calculation for an order (used when calculation_id is missing)"""
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        from sqlalchemy import update
        from database import calculations, orders, save_manual_edit, uploads, periods, save_calculation
        
        data = await request.json()
        fuel_payment = data.get("fuel_payment", 0)
        transport = data.get("transport", 0)
        total = data.get("total", 0)
        
        # Get order info
        order_query = orders.select().where(orders.c.id == order_id)
        order_row = await database.fetch_one(order_query)
        
        if not order_row:
            raise HTTPException(status_code=404, detail="Order not found")
        
        upload_id = order_row["upload_id"]
        worker = order_row["worker"]
        
        # Check if calculation exists
        calc_query = calculations.select().where(calculations.c.order_id == order_id)
        calc_row = await database.fetch_one(calc_query)
        
        if calc_row:
            # Update existing calculation
            calc_id = calc_row["id"]
            update_values = {
                "fuel_payment": float(fuel_payment),
                "transport": float(transport),
                "total": float(total)
            }
            query = update(calculations).where(calculations.c.id == calc_id).values(**update_values)
            await database.execute(query)
            logger.info("âœ… Updated calculation for order {order_id}: {update_values}")
        else:
            # Create new calculation
            calc_data = {
                "worker": worker,
                "fuel_payment": float(fuel_payment),
                "transport": float(transport),
                "total": float(total)
            }
            calc_id = await save_calculation(upload_id, order_id, calc_data)
            logger.info("âœ… Created calculation {calc_id} for order {order_id}")
        
        return JSONResponse({
            "success": True,
            "calculation_id": calc_id
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/calculation/{calc_id}/update")
async def update_calculation(calc_id: int, request: Request):
    """Update calculation values (fuel, transport, total)"""
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        # Get current user from session (safely)
        user = None
        try:
            user = request.session.get("user")
        except Exception as e:
            if DEBUG_MODE:
                logger.debug(f"Session not available: {e}")
        
        data = await request.json()
        fuel_payment = data.get("fuel_payment")
        transport = data.get("transport")
        total = data.get("total")
        
        # Build update query
        from sqlalchemy import update
        from database import calculations, orders, save_manual_edit, uploads, periods
        
        # First, get current values to record the change
        calc_query = calculations.select().where(calculations.c.id == calc_id)
        calc_row = await database.fetch_one(calc_query)
        
        if not calc_row:
            raise HTTPException(status_code=404, detail="Calculation not found")
        
        # Get order info for logging
        order_query = orders.select().where(orders.c.id == calc_row["order_id"])
        order_row = await database.fetch_one(order_query)
        
        # Get period status
        upload_id = calc_row["upload_id"]
        upload_query = uploads.select().where(uploads.c.id == upload_id)
        upload_row = await database.fetch_one(upload_query)
        period_status = "DRAFT"
        if upload_row:
            period_query = periods.select().where(periods.c.id == upload_row["period_id"])
            period_row = await database.fetch_one(period_query)
            if period_row:
                period_status = period_row["status"] if period_row["status"] else "DRAFT"
        
        update_values = {}
        edits_to_save = []
        
        # Prepare updates and track changes
        field_names = {
            "fuel_payment": "Ð‘ÐµÐ½Ð·Ð¸Ð½",
            "transport": "Ð¢Ñ€Ð°Ð½ÑÐ¿Ð¾Ñ€Ñ‚Ð½Ñ‹Ðµ", 
            "total": "Ð˜Ñ‚Ð¾Ð³Ð¾"
        }
        
        for field, new_val in [("fuel_payment", fuel_payment), ("transport", transport), ("total", total)]:
            if new_val is not None:
                old_val = calc_row[field] or 0
                new_val_float = float(new_val)
                if old_val != new_val_float:
                    update_values[field] = new_val_float
                    edits_to_save.append({
                        "field": field,
                        "old_value": old_val,
                        "new_value": new_val_float
                    })
        
        if not update_values:
            return JSONResponse({"success": True, "updated": {}, "message": "No changes"})
        
        # Update calculation
        query = update(calculations).where(calculations.c.id == calc_id).values(**update_values)
        await database.execute(query)
        
        # Save manual edits to history
        upload_id = calc_row["upload_id"]
        order_code = order_row["order_code"] if order_row else ""
        worker = calc_row["worker"]
        # For extra_rows, address is stored in order_full
        address = order_row["address"] if order_row and order_row["address"] else ""
        if not address and order_row and order_row["order_full"]:
            address = order_row["order_full"][:200]  # Use order_full as address for extra rows
        
        for edit in edits_to_save:
            await save_manual_edit(
                upload_id=upload_id,
                order_id=calc_row["order_id"],
                calculation_id=calc_id,
                order_code=order_code,
                worker=worker,
                address=address,
                field_name=edit["field"],
                old_value=edit["old_value"],
                new_value=edit["new_value"],
                edited_by=user.get("id") if user else None,
                edited_by_name=user.get("name") if user else None,
                period_status=period_status
            )
            logger.info(f"ðŸ“ Manual edit saved: {order_code} {worker} - {edit['field']}: {edit['old_value']} â†’ {edit['new_value']} by {user.get('name') if user else 'Unknown'} (status: {period_status})")
        
        # Update worker_totals
        full_worker = calc_row["worker"]
        base_worker = full_worker.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
        
        from database import worker_totals
        from sqlalchemy import and_, text
        
        # Get all orders and calculations for this worker using JOIN
        # This correctly uses orders.is_client_payment flag
        query = text("""
            SELECT c.total, c.fuel_payment, c.transport, o.is_client_payment
            FROM calculations c
            JOIN orders o ON c.order_id = o.id
            WHERE c.upload_id = :upload_id
            AND (o.worker = :worker OR o.worker = :worker_client)
        """).bindparams(
            upload_id=upload_id,
            worker=base_worker,
            worker_client=f"{base_worker} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)"
        )
        
        all_calcs = await database.fetch_all(query)
        
        company_amount = sum(c["total"] or 0 for c in all_calcs if not c["is_client_payment"])
        client_amount = sum(c["total"] or 0 for c in all_calcs if c["is_client_payment"])
        total_amount = company_amount + client_amount
        
        new_fuel = sum(c["fuel_payment"] or 0 for c in all_calcs)
        new_transport = sum(c["transport"] or 0 for c in all_calcs)
        
        # Update worker_totals
        update_wt = update(worker_totals).where(
            and_(
                worker_totals.c.upload_id == upload_id,
                worker_totals.c.worker == base_worker
            )
        ).values(
            total_amount=total_amount,
            company_amount=company_amount,
            client_amount=client_amount,
            fuel_total=new_fuel,
            transport_total=new_transport
        )
        await database.execute(update_wt)
        
        logger.info("âœ… Updated calculation {calc_id}: {update_values}")
        if DEBUG_MODE: logger.debug(f"   Worker {base_worker}: company={company_amount}, client={client_amount}, total={total_amount}")
        
        return JSONResponse({
            "success": True,
            "updated": update_values
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/order/{order_id}")
async def delete_order(order_id: int, request: Request):
    """Delete an order and its calculation from the database"""
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        from sqlalchemy import delete, and_, update, text
        from database import orders, calculations, worker_totals, manual_edits, log_action, save_manual_edit
        
        # Get current user (safely)
        user = None
        try:
            user = request.session.get("user")
        except Exception as e:
            if DEBUG_MODE:
                logger.debug(f"Session not available: {e}")
        
        # First, get order info for updating worker_totals
        order_query = orders.select().where(orders.c.id == order_id)
        order = await database.fetch_one(order_query)
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        upload_id = order["upload_id"]
        full_worker = order["worker"]
        base_worker = full_worker.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
        order_code = order["order_code"] or ""
        address = order["address"] or order.get("order_full", "")[:100] if order.get("order_full") else ""
        
        # Get period status
        from database import uploads, periods
        upload_query = uploads.select().where(uploads.c.id == upload_id)
        upload_row = await database.fetch_one(upload_query)
        period_status = "DRAFT"
        if upload_row:
            period_query = periods.select().where(periods.c.id == upload_row["period_id"])
            period_row = await database.fetch_one(period_query)
            if period_row:
                period_status = period_row["status"] if period_row["status"] else "DRAFT"
        
        # Get calculation info for logging
        calc_query = calculations.select().where(calculations.c.order_id == order_id)
        calc = await database.fetch_one(calc_query)
        
        deleted_total = calc["total"] if calc else 0
        calc_id = calc["id"] if calc else None
        
        # Save deletion as manual_edit with special field_name "DELETED"
        if calc:
            await save_manual_edit(
                upload_id=upload_id,
                order_id=order_id,
                calculation_id=calc_id,
                order_code=order_code,
                worker=full_worker,
                address=address,
                field_name="DELETED",
                old_value=deleted_total,
                new_value=0,
                edited_by=user.get("id") if user else None,
                edited_by_name=user.get("name") if user else None,
                period_status=period_status
            )
            logger.info(f"ðŸ“ Deletion saved to history: {order_code} {full_worker} - total was {deleted_total} by {user.get('name') if user else 'Unknown'} (status: {period_status})")
        
        # Delete manual_edits first (they reference calculation)
        if calc_id:
            # Don't delete the DELETED record we just created
            del_edits = delete(manual_edits).where(
                and_(
                    manual_edits.c.calculation_id == calc_id,
                    manual_edits.c.field_name != "DELETED"
                )
            )
            await database.execute(del_edits)
        
        # Also delete manual_edits by order_id (but keep DELETED records for history)
        del_edits_order = delete(manual_edits).where(
            and_(
                manual_edits.c.order_id == order_id,
                manual_edits.c.field_name != "DELETED"
            )
        )
        await database.execute(del_edits_order)
        
        # Delete calculation (foreign key to orders)
        del_calc = delete(calculations).where(calculations.c.order_id == order_id)
        await database.execute(del_calc)
        
        # Delete order
        del_order = delete(orders).where(orders.c.id == order_id)
        await database.execute(del_order)
        
        # FULL RECALCULATION of worker_totals (same logic as update_calculation)
        # This is more reliable than incremental subtraction
        query = text("""
            SELECT c.total, c.fuel_payment, c.transport, o.is_client_payment
            FROM calculations c
            JOIN orders o ON c.order_id = o.id
            WHERE c.upload_id = :upload_id
            AND (o.worker = :worker OR o.worker = :worker_client)
        """).bindparams(
            upload_id=upload_id,
            worker=base_worker,
            worker_client=f"{base_worker} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)"
        )
        
        all_calcs = await database.fetch_all(query)
        
        company_amount = sum(c["total"] or 0 for c in all_calcs if not c["is_client_payment"])
        client_amount = sum(c["total"] or 0 for c in all_calcs if c["is_client_payment"])
        total_amount = company_amount + client_amount
        
        new_fuel = sum(c["fuel_payment"] or 0 for c in all_calcs)
        new_transport = sum(c["transport"] or 0 for c in all_calcs)
        
        # Count orders
        company_count = sum(1 for c in all_calcs if not c["is_client_payment"])
        client_count = sum(1 for c in all_calcs if c["is_client_payment"])
        
        # Update worker_totals
        update_wt = update(worker_totals).where(
            and_(
                worker_totals.c.upload_id == upload_id,
                worker_totals.c.worker == base_worker
            )
        ).values(
            total_amount=total_amount,
            company_amount=company_amount,
            client_amount=client_amount,
            fuel_total=new_fuel,
            transport_total=new_transport,
            orders_count=company_count + client_count,
            company_orders_count=company_count,
            client_orders_count=client_count
        )
        await database.execute(update_wt)
        
        logger.info(f"ðŸ—‘ï¸ Deleted order {order_id} (worker: {base_worker}, deleted_total: {deleted_total})")
        if DEBUG_MODE: logger.debug(f"   Recalculated: company={company_amount}, client={client_amount}, total={total_amount}")
        
        return JSONResponse({
            "success": True,
            "deleted_order_id": order_id,
            "deleted_total": deleted_total
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload/{upload_id}/worker/{worker}/add-row")
async def add_order_row(upload_id: int, worker: str, request: Request):
    """Add a new order row for a worker"""
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        from urllib.parse import unquote
        from sqlalchemy import and_, update
        from database import orders, calculations, worker_totals
        
        worker_decoded = unquote(worker)
        data = await request.json()
        
        order_code = data.get("order_code", "")
        address = data.get("address", "")
        fuel_payment = float(data.get("fuel_payment", 0) or 0)
        transport = float(data.get("transport", 0) or 0)
        total = float(data.get("total", 0) or 0)
        
        # Determine if client payment
        is_client_payment = "(Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)" in worker_decoded
        base_worker = worker_decoded.replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "")
        
        # Create order text
        order_text = f"{order_code}, {address}" if order_code and address else (order_code or address or "Ð ÑƒÑ‡Ð½Ð°Ñ Ð·Ð°Ð¿Ð¸ÑÑŒ")
        
        # Insert new order
        order_insert = orders.insert().values(
            upload_id=upload_id,
            worker=worker_decoded,
            order_code=order_code,
            order_full=order_text,
            address=address,
            revenue_total=0,
            revenue_services=0,
            diagnostic=0,
            diagnostic_payment=0,
            specialist_fee=0,
            additional_expenses=0,
            service_payment=0,
            percent="",
            is_client_payment=is_client_payment,
            is_over_10k=False,
            is_extra_row=True  # Mark as manual/extra row
        )
        order_id = await database.execute(order_insert)
        
        # Insert calculation
        calc_insert = calculations.insert().values(
            upload_id=upload_id,
            order_id=order_id,
            worker=worker_decoded,
            fuel_payment=fuel_payment,
            transport=transport,
            diagnostic_50=0,
            total=total
        )
        calc_id = await database.execute(calc_insert)
        
        # FULL RECALCULATION of worker_totals (same logic as update_calculation)
        from sqlalchemy import text
        
        query = text("""
            SELECT c.total, c.fuel_payment, c.transport, o.is_client_payment
            FROM calculations c
            JOIN orders o ON c.order_id = o.id
            WHERE c.upload_id = :upload_id
            AND (o.worker = :worker OR o.worker = :worker_client)
        """).bindparams(
            upload_id=upload_id,
            worker=base_worker,
            worker_client=f"{base_worker} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)"
        )
        
        all_calcs = await database.fetch_all(query)
        
        company_amount = sum(c["total"] or 0 for c in all_calcs if not c["is_client_payment"])
        client_amount = sum(c["total"] or 0 for c in all_calcs if c["is_client_payment"])
        total_amount = company_amount + client_amount
        
        new_fuel = sum(c["fuel_payment"] or 0 for c in all_calcs)
        new_transport = sum(c["transport"] or 0 for c in all_calcs)
        
        # Count orders
        company_count = sum(1 for c in all_calcs if not c["is_client_payment"])
        client_count = sum(1 for c in all_calcs if c["is_client_payment"])
        
        # Update worker_totals
        update_wt = update(worker_totals).where(
            and_(
                worker_totals.c.upload_id == upload_id,
                worker_totals.c.worker == base_worker
            )
        ).values(
            total_amount=total_amount,
            company_amount=company_amount,
            client_amount=client_amount,
            fuel_total=new_fuel,
            transport_total=new_transport,
            orders_count=company_count + client_count,
            company_orders_count=company_count,
            client_orders_count=client_count
        )
        await database.execute(update_wt)
        
        logger.info(f"âž• Added new row for {worker_decoded}: order_code={order_code}, total={total}")
        if DEBUG_MODE: logger.debug(f"   Recalculated: company={company_amount}, client={client_amount}, total={total_amount}")
        
        # Save to manual_edits for history tracking
        from database import save_manual_edit, uploads, periods
        
        # Get period status
        upload_query = uploads.select().where(uploads.c.id == upload_id)
        upload_row = await database.fetch_one(upload_query)
        period_status = "DRAFT"
        if upload_row:
            period_query = periods.select().where(periods.c.id == upload_row["period_id"])
            period_row = await database.fetch_one(period_query)
            if period_row:
                period_status = period_row["status"] if period_row["status"] else "DRAFT"
        
        # Save as "ADDED" manual edit
        await save_manual_edit(
            upload_id=upload_id,
            order_id=order_id,
            calculation_id=calc_id,
            order_code=order_code or address or "Ð ÑƒÑ‡Ð½Ð°Ñ Ð·Ð°Ð¿Ð¸ÑÑŒ",
            worker=base_worker,
            address=address,
            field_name="ADDED",
            old_value=0,
            new_value=total,
            period_status=period_status
        )
        logger.info(f"ðŸ“ Saved manual edit for new row: {order_code or address}")
        
        # Return the new order data
        return JSONResponse({
            "success": True,
            "order": {
                "id": order_id,
                "order_code": order_code,
                "address": address,
                "order_full": order_text,
                "worker": worker_decoded,
                "revenue_services": 0,
                "service_payment": 0,
                "percent": "",
                "is_client_payment": is_client_payment,
                "is_over_10k": False,
                "is_extra_row": True,
                "calculation": {
                    "id": calc_id,
                    "fuel_payment": fuel_payment,
                    "transport": transport,
                    "total": total
                }
            }
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/order/{order_id}/update")
async def update_order_info(order_id: int, request: Request):
    """Update order info (order_code, address) - for manually added rows"""
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        from sqlalchemy import update
        from database import orders
        
        data = await request.json()
        
        update_values = {}
        if "order_code" in data:
            update_values["order_code"] = data["order_code"]
        if "address" in data:
            update_values["address"] = data["address"]
        
        if update_values:
            # Also update order_full
            order_query = orders.select().where(orders.c.id == order_id)
            order = await database.fetch_one(order_query)
            
            if order:
                new_order_code = data.get("order_code", order["order_code"] or "")
                new_address = data.get("address", order["address"] or "")
                order_text = f"{new_order_code}, {new_address}" if new_order_code and new_address else (new_order_code or new_address or "Ð ÑƒÑ‡Ð½Ð°Ñ Ð·Ð°Ð¿Ð¸ÑÑŒ")
                update_values["order_full"] = order_text
        
        if not update_values:
            return JSONResponse({"success": True, "message": "No changes"})
        
        query = update(orders).where(orders.c.id == order_id).values(**update_values)
        await database.execute(query)
        
        logger.info(f"ðŸ“ Updated order {order_id}: {update_values}")
        
        return JSONResponse({
            "success": True,
            "updated": update_values
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/period/{period_id}")
async def period_page(request: Request, period_id: int):
    """Period details page"""
    user = get_current_user(request)
    response = templates.TemplateResponse("period.html", {"request": request, "period_id": period_id, "user": user})
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/upload/{upload_id}")
async def upload_page(request: Request, upload_id: int, worker: str = ""):
    """Upload details page"""
    # Get worker from query parameter
    worker_name = worker or request.query_params.get("worker", "")
    user = get_current_user(request)
    response = templates.TemplateResponse("upload.html", {
        "request": request, 
        "upload_id": upload_id,
        "worker_name": worker_name,
        "user": user
    })
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# ============================================================================
# 1C INTEGRATION
# ============================================================================

# 1C Configuration - UPDATE THESE VALUES
ONEÐ¡_CONFIG = {
    "enabled": False,  # Set to True when 1C HTTP service is ready
    "base_url": "http://your-1c-server/your-database/hs/salary",  # Update with your 1C server URL
    "username": "api_user",  # 1C username for API access
    "password": "api_password",  # 1C password
    "timeout": 10,  # Request timeout in seconds
}


@app.get("/api/1c/order/{order_code}")
async def get_1c_order_info(order_code: str):
    """
    Get order information from 1C.
    This endpoint proxies requests to the 1C HTTP service.
    """
    import httpx
    from urllib.parse import quote
    
    # Check if 1C integration is enabled
    if not ONEÐ¡_CONFIG["enabled"]:
        return JSONResponse({
            "success": False,
            "error": "Ð˜Ð½Ñ‚ÐµÐ³Ñ€Ð°Ñ†Ð¸Ñ Ñ 1Ð¡ Ð½Ðµ Ð½Ð°ÑÑ‚Ñ€Ð¾ÐµÐ½Ð°",
            "hint": "ÐÐµÐ¾Ð±Ñ…Ð¾Ð´Ð¸Ð¼Ð¾ Ð½Ð°ÑÑ‚Ñ€Ð¾Ð¸Ñ‚ÑŒ HTTP-ÑÐµÑ€Ð²Ð¸Ñ Ð² 1Ð¡ Ð¸ Ð¾Ð±Ð½Ð¾Ð²Ð¸Ñ‚ÑŒ ÐºÐ¾Ð½Ñ„Ð¸Ð³ÑƒÑ€Ð°Ñ†Ð¸ÑŽ"
        })
    
    try:
        # Build 1C API URL
        url = f"{ONEÐ¡_CONFIG['base_url']}/orders/{quote(order_code)}"
        
        # Make request to 1C with basic auth
        auth = (ONEÐ¡_CONFIG["username"], ONEÐ¡_CONFIG["password"])
        
        async with httpx.AsyncClient(timeout=ONEÐ¡_CONFIG["timeout"]) as client:
            response = await client.get(url, auth=auth)
            
            if response.status_code == 200:
                data = response.json()
                return JSONResponse(data)
            elif response.status_code == 401:
                return JSONResponse({
                    "success": False,
                    "error": "ÐžÑˆÐ¸Ð±ÐºÐ° Ð°Ð²Ñ‚Ð¾Ñ€Ð¸Ð·Ð°Ñ†Ð¸Ð¸ Ð² 1Ð¡"
                })
            elif response.status_code == 404:
                return JSONResponse({
                    "success": False,
                    "error": f"Ð—Ð°ÐºÐ°Ð· {order_code} Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½ Ð² 1Ð¡"
                })
            else:
                return JSONResponse({
                    "success": False,
                    "error": f"ÐžÑˆÐ¸Ð±ÐºÐ° 1Ð¡: {response.status_code}"
                })
                
    except httpx.TimeoutException:
        return JSONResponse({
            "success": False,
            "error": "ÐŸÑ€ÐµÐ²Ñ‹ÑˆÐµÐ½Ð¾ Ð²Ñ€ÐµÐ¼Ñ Ð¾Ð¶Ð¸Ð´Ð°Ð½Ð¸Ñ Ð¾Ñ‚Ð²ÐµÑ‚Ð° Ð¾Ñ‚ 1Ð¡"
        })
    except httpx.ConnectError:
        return JSONResponse({
            "success": False,
            "error": "ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ Ð¿Ð¾Ð´ÐºÐ»ÑŽÑ‡Ð¸Ñ‚ÑŒÑÑ Ðº ÑÐµÑ€Ð²ÐµÑ€Ñƒ 1Ð¡"
        })
    except Exception as e:
        return JSONResponse({
            "success": False,
            "error": f"ÐžÑˆÐ¸Ð±ÐºÐ°: {str(e)}"
        })


@app.get("/api/1c/status")
async def get_1c_status():
    """Check 1C integration status"""
    return JSONResponse({
        "enabled": ONEÐ¡_CONFIG["enabled"],
        "base_url": ONEÐ¡_CONFIG["base_url"] if ONEÐ¡_CONFIG["enabled"] else None,
        "message": "Ð˜Ð½Ñ‚ÐµÐ³Ñ€Ð°Ñ†Ð¸Ñ Ñ 1Ð¡ Ð°ÐºÑ‚Ð¸Ð²Ð½Ð°" if ONEÐ¡_CONFIG["enabled"] else "Ð˜Ð½Ñ‚ÐµÐ³Ñ€Ð°Ñ†Ð¸Ñ Ñ 1Ð¡ Ð½Ðµ Ð½Ð°ÑÑ‚Ñ€Ð¾ÐµÐ½Ð°"
    })


@app.post("/api/upload/{upload_id}/recalculate")
async def recalculate_worker_totals(upload_id: int):
    """
    Force recalculation of all worker_totals for an upload.
    Use this to fix any inconsistencies in the data.
    """
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        from sqlalchemy import and_, update, text
        from database import worker_totals
        
        # Get all workers for this upload
        wt_query = worker_totals.select().where(worker_totals.c.upload_id == upload_id)
        all_wt = await database.fetch_all(wt_query)
        
        recalculated = []
        
        for wt in all_wt:
            base_worker = wt["worker"]
            
            # Full recalculation using JOIN
            query = text("""
                SELECT c.total, c.fuel_payment, c.transport, o.is_client_payment
                FROM calculations c
                JOIN orders o ON c.order_id = o.id
                WHERE c.upload_id = :upload_id
                AND (o.worker = :worker OR o.worker = :worker_client)
            """).bindparams(
                upload_id=upload_id,
                worker=base_worker,
                worker_client=f"{base_worker} (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)"
            )
            
            all_calcs = await database.fetch_all(query)
            
            company_amount = sum(c["total"] or 0 for c in all_calcs if not c["is_client_payment"])
            client_amount = sum(c["total"] or 0 for c in all_calcs if c["is_client_payment"])
            total_amount = company_amount + client_amount
            
            new_fuel = sum(c["fuel_payment"] or 0 for c in all_calcs)
            new_transport = sum(c["transport"] or 0 for c in all_calcs)
            
            company_count = sum(1 for c in all_calcs if not c["is_client_payment"])
            client_count = sum(1 for c in all_calcs if c["is_client_payment"])
            
            # Update worker_totals
            update_wt = update(worker_totals).where(
                and_(
                    worker_totals.c.upload_id == upload_id,
                    worker_totals.c.worker == base_worker
                )
            ).values(
                total_amount=total_amount,
                company_amount=company_amount,
                client_amount=client_amount,
                fuel_total=new_fuel,
                transport_total=new_transport,
                orders_count=company_count + client_count,
                company_orders_count=company_count,
                client_orders_count=client_count
            )
            await database.execute(update_wt)
            
            recalculated.append({
                "worker": base_worker,
                "company_amount": company_amount,
                "client_amount": client_amount,
                "total_amount": total_amount
            })
            
            if DEBUG_MODE: logger.debug(f"ðŸ”„ Recalculated {base_worker}: company={company_amount}, client={client_amount}, total={total_amount}")
        
        return JSONResponse({
            "success": True,
            "recalculated_count": len(recalculated),
            "workers": recalculated
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/recalculate-all-totals")
async def recalculate_all_totals():
    """Recalculate worker_totals for ALL uploads based on actual order calculations"""
    try:
        from sqlalchemy import text
        
        # Get all uploads
        query = text("SELECT DISTINCT upload_id FROM orders")
        uploads = await database.fetch_all(query)
        
        recalculated_uploads = []
        
        for upload_row in uploads:
            upload_id = upload_row["upload_id"]
            
            # Get all orders with calculations for this upload
            query = text("""
                SELECT o.worker, o.is_client_payment, c.total
                FROM orders o
                JOIN calculations c ON c.order_id = o.id
                WHERE o.upload_id = :upload_id
            """).bindparams(upload_id=upload_id)
            
            order_calcs = await database.fetch_all(query)
            
            # Aggregate by worker
            worker_sums = {}
            for oc in order_calcs:
                worker = oc["worker"]
                if worker not in worker_sums:
                    worker_sums[worker] = {"company": 0, "client": 0}
                
                total = oc["total"] or 0
                if oc["is_client_payment"]:
                    worker_sums[worker]["client"] += total
                else:
                    worker_sums[worker]["company"] += total
            
            # Update worker_totals
            for worker, sums in worker_sums.items():
                # Check if worker_total exists
                check_query = text("""
                    SELECT id FROM worker_totals WHERE upload_id = :upload_id AND worker = :worker
                """).bindparams(upload_id=upload_id, worker=worker)
                existing = await database.fetch_one(check_query)
                
                if existing:
                    update_query = text("""
                        UPDATE worker_totals 
                        SET company_amount = :company, client_amount = :client, total_amount = :total
                        WHERE upload_id = :upload_id AND worker = :worker
                    """).bindparams(
                        company=sums["company"],
                        client=sums["client"],
                        total=sums["company"] + sums["client"],
                        upload_id=upload_id,
                        worker=worker
                    )
                    await database.execute(update_query)
                else:
                    insert_query = text("""
                        INSERT INTO worker_totals (upload_id, worker, company_amount, client_amount, total_amount)
                        VALUES (:upload_id, :worker, :company, :client, :total)
                    """).bindparams(
                        upload_id=upload_id,
                        worker=worker,
                        company=sums["company"],
                        client=sums["client"],
                        total=sums["company"] + sums["client"]
                    )
                    await database.execute(insert_query)
            
            recalculated_uploads.append({
                "upload_id": upload_id,
                "workers_count": len(worker_sums)
            })
            logger.info("âœ… Recalculated upload {upload_id}: {len(worker_sums)} workers")
        
        return JSONResponse({
            "success": True,
            "recalculated_uploads": len(recalculated_uploads),
            "details": recalculated_uploads
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/delete-upload/{upload_id}")
async def delete_upload(upload_id: int):
    """Delete an upload and all its related data (GET for browser access)"""
    from sqlalchemy import text
    try:
        # Delete in correct order due to foreign keys
        await database.execute(text("DELETE FROM manual_edits WHERE upload_id = :id").bindparams(id=upload_id))
        await database.execute(text("DELETE FROM changes WHERE upload_id = :id").bindparams(id=upload_id))
        await database.execute(text("DELETE FROM worker_totals WHERE upload_id = :id").bindparams(id=upload_id))
        await database.execute(text("DELETE FROM calculations WHERE upload_id = :id").bindparams(id=upload_id))
        await database.execute(text("DELETE FROM orders WHERE upload_id = :id").bindparams(id=upload_id))
        await database.execute(text("DELETE FROM uploads WHERE id = :id").bindparams(id=upload_id))
        
        logger.info("âœ… Deleted upload {upload_id}")
        return JSONResponse({"success": True, "deleted_upload_id": upload_id})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/api/list-uploads/{period_id}")
async def list_uploads(period_id: int):
    """List all uploads for a period with their order counts"""
    from sqlalchemy import text
    try:
        query = text("""
            SELECT u.id, u.version, u.created_at,
                   (SELECT COUNT(*) FROM orders WHERE upload_id = u.id) as orders_count
            FROM uploads u
            WHERE u.period_id = :period_id
            ORDER BY u.version DESC
        """)
        results = await database.fetch_all(query.bindparams(period_id=period_id))
        
        uploads = []
        for r in results:
            uploads.append({
                "id": r["id"],
                "version": r["version"],
                "created_at": str(r["created_at"]),
                "orders_count": r["orders_count"]
            })
        
        return JSONResponse({"success": True, "uploads": uploads})
        
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})


@app.delete("/api/period/{period_id}")
async def delete_period(period_id: int, request: Request):
    """Delete a period and all related data (admin only)"""
    try:
        if not database:
            raise HTTPException(status_code=500, detail="Database not connected")
        
        # Check admin permission
        user = get_current_user(request)
        
        if not user or user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Ð¢Ð¾Ð»ÑŒÐºÐ¾ Ð°Ð´Ð¼Ð¸Ð½Ð¸ÑÑ‚Ñ€Ð°Ñ‚Ð¾Ñ€ Ð¼Ð¾Ð¶ÐµÑ‚ ÑƒÐ´Ð°Ð»ÑÑ‚ÑŒ Ð¿ÐµÑ€Ð¸Ð¾Ð´Ñ‹")
        
        from sqlalchemy import delete
        from database import periods, uploads, orders, calculations, worker_totals, manual_edits, changes, audit_log
        
        # Check period exists
        period_query = periods.select().where(periods.c.id == period_id)
        period = await database.fetch_one(period_query)
        
        if not period:
            raise HTTPException(status_code=404, detail="ÐŸÐµÑ€Ð¸Ð¾Ð´ Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½")
        
        period_name = period["name"]
        
        # Get all upload_ids for this period
        uploads_query = uploads.select().where(uploads.c.period_id == period_id)
        period_uploads = await database.fetch_all(uploads_query)
        upload_ids = [u["id"] for u in period_uploads]
        
        if upload_ids:
            # Delete in correct order (foreign key constraints)
            # 1. manual_edits
            await database.execute(
                delete(manual_edits).where(manual_edits.c.upload_id.in_(upload_ids))
            )
            
            # 2. changes
            await database.execute(
                delete(changes).where(changes.c.upload_id.in_(upload_ids))
            )
            
            # 3. calculations
            await database.execute(
                delete(calculations).where(calculations.c.upload_id.in_(upload_ids))
            )
            
            # 4. orders
            await database.execute(
                delete(orders).where(orders.c.upload_id.in_(upload_ids))
            )
            
            # 5. worker_totals
            await database.execute(
                delete(worker_totals).where(worker_totals.c.upload_id.in_(upload_ids))
            )
            
            # 6. uploads
            await database.execute(
                delete(uploads).where(uploads.c.period_id == period_id)
            )
        
        # 7. period itself
        await database.execute(
            delete(periods).where(periods.c.id == period_id)
        )
        
        logger.info(f"ðŸ—‘ï¸ Period '{period_name}' (id={period_id}) deleted by {user.get('name', 'Unknown')}")
        
        return JSONResponse({
            "success": True,
            "message": f"ÐŸÐµÑ€Ð¸Ð¾Ð´ '{period_name}' ÑƒÐ´Ð°Ð»Ñ‘Ð½"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============== SEARCH ==============

@app.get("/api/search")
async def search_orders(q: str = "", limit: int = 10):
    """Search orders by order_code, address, worker, or amount with fuzzy matching"""
    try:
        if not database or not database.is_connected:
            return JSONResponse({"success": False, "error": "Database not connected"})
        
        if not q or len(q) < 2:
            return JSONResponse({"success": True, "results": []})
        
        # Normalize query: replace Ñ‘ with Ðµ for consistent matching
        q_normalized = q.replace('Ñ‘', 'Ðµ').replace('Ð', 'Ð•')
        search_term = f"%{q_normalized}%"
        
        # Also create version with Ðµ replaced by Ñ‘ for reverse matching
        q_with_yo = q.replace('Ðµ', 'Ñ‘').replace('Ð•', 'Ð')
        search_term_yo = f"%{q_with_yo}%"
        
        # Try to parse as number for amount search
        try:
            amount_search = float(q.replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            amount_search = None
        
        # First, try to enable pg_trgm extension for fuzzy search (ignore if fails)
        try:
            await database.execute("CREATE EXTENSION IF NOT EXISTS pg_trgm")
        except Exception as e:
            if DEBUG_MODE:
                logger.debug(f"pg_trgm extension not available: {e}")
        
        # Build query with fuzzy matching using ILIKE and similarity
        # Uses trigram similarity for typo tolerance
        query = """
            SELECT 
                o.id as order_id,
                o.order_code,
                o.order_date,
                o.address,
                o.worker,
                o.revenue_services,
                o.service_payment,
                o.percent,
                o.is_client_payment,
                o.manager_comment,
                c.fuel_payment,
                c.transport,
                c.total,
                p.id as period_id,
                p.name as period_name,
                u.id as upload_id,
                GREATEST(
                    COALESCE(similarity(LOWER(REPLACE(o.order_code, 'Ñ‘', 'Ðµ')), LOWER(:q_norm)), 0),
                    COALESCE(similarity(LOWER(REPLACE(o.address, 'Ñ‘', 'Ðµ')), LOWER(:q_norm)), 0),
                    COALESCE(similarity(LOWER(REPLACE(o.worker, 'Ñ‘', 'Ðµ')), LOWER(:q_norm)), 0)
                ) as match_score
            FROM orders o
            LEFT JOIN calculations c ON o.id = c.order_id
            INNER JOIN uploads u ON o.upload_id = u.id
            INNER JOIN (
                SELECT period_id, MAX(id) as latest_upload_id
                FROM uploads
                GROUP BY period_id
            ) lu ON u.id = lu.latest_upload_id
            LEFT JOIN periods p ON u.period_id = p.id
            WHERE (
                -- Exact/partial match with Ñ‘â†’Ðµ normalization
                REPLACE(LOWER(o.order_code), 'Ñ‘', 'Ðµ') ILIKE LOWER(:search)
                OR REPLACE(LOWER(o.address), 'Ñ‘', 'Ðµ') ILIKE LOWER(:search)
                OR REPLACE(LOWER(o.worker), 'Ñ‘', 'Ðµ') ILIKE LOWER(:search)
                -- Also try with Ðµâ†’Ñ‘
                OR LOWER(o.order_code) ILIKE LOWER(:search_yo)
                OR LOWER(o.address) ILIKE LOWER(:search_yo)
                OR LOWER(o.worker) ILIKE LOWER(:search_yo)
                -- Amount search
                OR (c.total = :amount AND :amount IS NOT NULL)
                OR (o.service_payment = :amount AND :amount IS NOT NULL)
                OR (o.revenue_services = :amount AND :amount IS NOT NULL)
                -- Fuzzy matching for typos (similarity > 0.3)
                OR similarity(LOWER(REPLACE(o.order_code, 'Ñ‘', 'Ðµ')), LOWER(:q_norm)) > 0.3
                OR similarity(LOWER(REPLACE(o.address, 'Ñ‘', 'Ðµ')), LOWER(:q_norm)) > 0.3
                OR similarity(LOWER(REPLACE(o.worker, 'Ñ‘', 'Ðµ')), LOWER(:q_norm)) > 0.3
            )
            ORDER BY match_score DESC, p.created_at DESC, o.id DESC
            LIMIT :limit
        """
        
        try:
            rows = await database.fetch_all(query, {
                "search": search_term,
                "search_yo": search_term_yo,
                "q_norm": q_normalized,
                "amount": amount_search,
                "limit": limit
            })
        except Exception as e:
            # Fallback to simple search if pg_trgm not available
            logger.warning("âš ï¸ Fuzzy search failed, using simple search: {e}")
            query = """
                SELECT 
                    o.id as order_id,
                    o.order_code,
                    o.order_date,
                    o.address,
                    o.worker,
                    o.revenue_services,
                    o.service_payment,
                    o.percent,
                    o.is_client_payment,
                    o.manager_comment,
                    c.fuel_payment,
                    c.transport,
                    c.total,
                    p.id as period_id,
                    p.name as period_name,
                    u.id as upload_id
                FROM orders o
                LEFT JOIN calculations c ON o.id = c.order_id
                INNER JOIN uploads u ON o.upload_id = u.id
                INNER JOIN (
                    SELECT period_id, MAX(id) as latest_upload_id
                    FROM uploads
                    GROUP BY period_id
                ) lu ON u.id = lu.latest_upload_id
                LEFT JOIN periods p ON u.period_id = p.id
                WHERE (
                    REPLACE(LOWER(o.order_code), 'Ñ‘', 'Ðµ') ILIKE LOWER(:search)
                    OR REPLACE(LOWER(o.address), 'Ñ‘', 'Ðµ') ILIKE LOWER(:search)
                    OR REPLACE(LOWER(o.worker), 'Ñ‘', 'Ðµ') ILIKE LOWER(:search)
                    OR LOWER(o.order_code) ILIKE LOWER(:search_yo)
                    OR LOWER(o.address) ILIKE LOWER(:search_yo)
                    OR LOWER(o.worker) ILIKE LOWER(:search_yo)
                    OR (c.total = :amount AND :amount IS NOT NULL)
                    OR (o.service_payment = :amount AND :amount IS NOT NULL)
                )
                ORDER BY p.created_at DESC, o.id DESC
                LIMIT :limit
            """
            rows = await database.fetch_all(query, {
                "search": search_term,
                "search_yo": search_term_yo,
                "amount": amount_search,
                "limit": limit
            })
        
        results = []
        for row in rows:
            r = dict(row._mapping)
            # Clean worker name (remove " (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)" suffix for display)
            worker_display = r["worker"].replace(" (Ð¾Ð¿Ð»Ð°Ñ‚Ð° ÐºÐ»Ð¸ÐµÐ½Ñ‚Ð¾Ð¼)", "") if r["worker"] else ""
            # Format order date
            order_date = r.get("order_date")
            order_date_str = order_date.strftime("%d.%m.%y") if order_date else ""
            results.append({
                "order_id": r["order_id"],
                "order_code": r["order_code"] or "-",
                "order_date": order_date_str,
                "address": r["address"] or "-",
                "worker": worker_display,
                "revenue": r["revenue_services"] or 0,
                "payment": r["service_payment"] or 0,
                "percent": r["percent"] or "-",
                "fuel": r["fuel_payment"] or 0,
                "transport": r["transport"] or 0,
                "total": r["total"] or 0,
                "type": "ÐšÐ»Ð¸ÐµÐ½Ñ‚" if r["is_client_payment"] else "ÐšÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ñ",
                "manager_comment": r["manager_comment"],
                "period_id": r["period_id"],
                "period_name": r["period_name"] or "-",
                "upload_id": r["upload_id"]
            })
        
        return JSONResponse({"success": True, "results": results, "query": q})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"success": False, "error": str(e)})


@app.get("/search")
async def search_page(request: Request, q: str = ""):
    """Search results page"""
    user = get_current_user(request)
    return templates.TemplateResponse("search.html", {
        "request": request,
        "user": user,
        "query": q,
        "csrf_token": request.state.csrf_token if hasattr(request.state, 'csrf_token') else ''
    })


# ===== DUPLICATE CHECK (Admin only) =====

@app.get("/api/duplicates")
async def get_duplicates(request: Request):
    """Get duplicate orders analysis (admin only)"""
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        return {"success": False, "error": "Admin access required"}
    
    try:
        import re
        from collections import defaultdict
        
        # Get latest upload for each period
        latest_uploads_query = """
            SELECT u.id as upload_id, u.period_id, p.name as period_name
            FROM uploads u
            JOIN periods p ON u.period_id = p.id
            WHERE u.version = (
                SELECT MAX(u2.version) FROM uploads u2 WHERE u2.period_id = u.period_id
            )
            ORDER BY p.name DESC
        """
        latest_uploads = await database.fetch_all(latest_uploads_query)
        upload_ids = [u._mapping["upload_id"] for u in latest_uploads]
        
        if not upload_ids:
            return {"success": True, "exact_duplicates": [], "partial_duplicates": [], "needs_review": [], "stats": {}}
        
        # Get all orders from latest uploads
        orders_query = f"""
            SELECT 
                o.id, o.upload_id, o.order_code, o.address, o.worker, o.order_full,
                o.order_date,
                o.revenue_total, o.revenue_services, o.diagnostic,
                o.is_client_payment,
                c.total,
                u.period_id, p.name as period_name
            FROM orders o
            JOIN uploads u ON o.upload_id = u.id
            JOIN periods p ON u.period_id = p.id
            LEFT JOIN calculations c ON c.order_id = o.id
            WHERE o.upload_id IN ({','.join(map(str, upload_ids))})
            ORDER BY p.name, o.worker
        """
        all_orders = await database.fetch_all(orders_query)
        
        # === HELPER FUNCTIONS ===
        
        def get_work_type(order):
            """Determine work type from order data"""
            diag = order._mapping.get("diagnostic") or 0
            revenue = order._mapping.get("revenue_services") or 0
            total = order._mapping.get("total") or 0
            
            order_full = (order._mapping.get("order_full") or "").lower()
            address = (order._mapping.get("address") or "").lower()
            combined = order_full + " " + address
            
            if diag > 0:
                return "diagnostic"
            if "Ð´Ð¸Ð°Ð³Ð½Ð¾ÑÑ‚Ð¸Ðº" in combined:
                return "diagnostic"
            if "Ð¾ÑÐ¼Ð¾Ñ‚Ñ€" in combined or "Ð²Ñ‹ÐµÐ·Ð´" in combined:
                return "inspection"
            if "Ð¼Ð¾Ð½Ñ‚Ð°Ð¶" in combined:
                return "installation"
            if revenue > 0 and total <= 5000:
                return "inspection"
            if total > 5000:
                return "installation"
            return "other"
        
        def normalize_address(addr):
            """Normalize address for comparison"""
            if not addr:
                return ""
            
            addr = addr.lower().strip()
            
            # Remove specific prefixes first
            addr = re.sub(r'^Ð¼Ð¾Ð½Ñ‚Ð°Ð¶!\s*', '', addr, flags=re.IGNORECASE)
            addr = re.sub(r'^ÑÐ´ÐµÐ»ÐºÐ°:\s*', '', addr, flags=re.IGNORECASE)
            
            remove_patterns = [
                r'Ð¼Ð¾ÑÐºÐ¾Ð²ÑÐºÐ°Ñ\s+Ð¾Ð±Ð»Ð°ÑÑ‚ÑŒ,?\s*',
                r'Ð³Ð¾Ñ€Ð¾Ð´ÑÐºÐ¾Ð¹\s+Ð¾ÐºÑ€ÑƒÐ³\s*[^,]*,?\s*',  # Remove "Ð³Ð¾Ñ€Ð¾Ð´ÑÐºÐ¾Ð¹ Ð¾ÐºÑ€ÑƒÐ³ Ð›ÑŽÐ±ÐµÑ€Ñ†Ñ‹" completely
                r'Ð¼ÑƒÐ½Ð¸Ñ†Ð¸Ð¿Ð°Ð»ÑŒÐ½Ñ‹Ð¹\s+Ð¾ÐºÑ€ÑƒÐ³\s*[^,]*,?\s*',
                r'Ð¿Ð¾ÑÑ‘Ð»Ð¾Ðº\s+Ð³Ð¾Ñ€Ð¾Ð´ÑÐºÐ¾Ð³Ð¾\s+Ñ‚Ð¸Ð¿Ð°\s+',
                r'Ð³\.?\s*Ð¾\.?\s*',
                r'Ð¼Ð¾ÑÐºÐ²Ð°,?\s*', r'Ð¼Ð¾,?\s*',
                r'Ð¿Ð¾Ñ\.\s*', r'Ð¿Ð¾ÑÑ‘Ð»Ð¾Ðº\s+', r'Ð¿Ð¾ÑÐµÐ»Ð¾Ðº\s+',
                r'Ð´ÐµÑ€ÐµÐ²Ð½Ñ\s+', r'Ð´ÐµÑ€\.\s*', r'ÑÐµÐ»Ð¾\s+', r'Ñ\.\s+',
                r'ÑƒÐ»Ð¸Ñ†Ð°\s+', r'ÑƒÐ»\.\s*', r'Ð¿Ñ€Ð¾ÑÐ¿ÐµÐºÑ‚\s+', r'Ð¿Ñ€-Ñ‚\.?\s*', r'Ð¿Ñ€\.\s*',
                r'ÑˆÐ¾ÑÑÐµ\s+', r'Ñˆ\.\s*', r'Ð¿ÐµÑ€ÐµÑƒÐ»Ð¾Ðº\s+', r'Ð¿ÐµÑ€\.\s*',
                r'Ñ€Ð°Ð¹Ð¾Ð½\s+', r'Ñ€-Ð½\s*', r'ÐºÐ¾Ñ€Ð¿ÑƒÑ\s+', r'ÐºÐ¾Ñ€Ð¿\.\s*', r'Ðº\.\s*',
                r'ÑÑ‚Ñ€Ð¾ÐµÐ½Ð¸Ðµ\s+', r'ÑÑ‚Ñ€\.\s*', r'Ð´Ð¾Ð¼\s+', r'Ð´\.\s*',
                r'ÑÐ½Ñ‚\s+', r'Ñ‚ÑÐ½\s+', r'ÐºÐ¿\s+', r'Ð´Ð½Ð¿\s+', r'Ð´Ðº\s+',
                r'Ð¾Ð¾Ð¾\s+[^,]+,?\s*', r'Ð´Ð»Ñ\s+Ð·Ð¿\s*', r'Â«|Â»|"|"',
                r'Ð»ÑŽÐ±ÐµÑ€ÐµÑ†ÐºÐ¸Ð¹\s*', r'Ð»ÑŽÐ±ÐµÑ€Ñ†Ñ‹\s*',
                r'-\s*ÑƒÐ»Ñ‹Ð±ÐºÐ°\s+Ñ€Ð°Ð´ÑƒÐ³Ð¸.*',
                r'\bÐ³\.\s*',
            ]
            
            for pattern in remove_patterns:
                addr = re.sub(pattern, ' ', addr, flags=re.IGNORECASE)
            
            addr = re.sub(r'[,.:;\\n]+', ' ', addr)
            addr = re.sub(r'\s+', ' ', addr)
            return addr.strip()
        
        def extract_street_and_number(addr):
            """Extract street name and house number for strict comparison"""
            norm = normalize_address(addr)
            
            # Find house number
            number_match = re.search(r'\b(\d+[Ð°-Ña-z]?)\b', norm)
            number = number_match.group(1) if number_match else ""
            
            # Remove numbers to get street name
            street = re.sub(r'\b\d+[Ð°-Ña-z]?\b', '', norm).strip()
            street = re.sub(r'\s+', ' ', street).strip()
            
            return street, number
        
        def addresses_match(addr1, addr2):
            """Strict address comparison - requires street AND number match"""
            if not addr1 or not addr2:
                return False
            
            s1, n1 = extract_street_and_number(addr1)
            s2, n2 = extract_street_and_number(addr2)
            
            # Numbers must match (if both exist)
            if n1 and n2 and n1 != n2:
                return False
            
            # Compare street words
            words1 = set(w for w in s1.split() if len(w) > 2)
            words2 = set(w for w in s2.split() if len(w) > 2)
            
            if not words1 or not words2:
                return False
            
            intersection = len(words1 & words2)
            union = len(words1 | words2)
            similarity = intersection / union if union > 0 else 0
            
            # Require 60% word overlap
            return similarity >= 0.5
        
        def get_address_from_order(order):
            """Extract address from order, including manual entries"""
            addr = order._mapping.get("address") or ""
            order_full = order._mapping.get("order_full") or ""
            
            if not addr and order_full:
                addr = order_full
            
            return addr
        
        def amounts_similar(a1, a2, tolerance=0.1):
            """Check if amounts are within tolerance (default 10%)"""
            if a1 == 0 and a2 == 0:
                return True
            if a1 == 0 or a2 == 0:
                return False
            diff = abs(a1 - a2) / max(a1, a2)
            return diff <= tolerance
        
        # === BUILD ORDER LIST ===
        
        processed_orders = []
        for o in all_orders:
            addr = get_address_from_order(o)
            processed_orders.append({
                "id": o._mapping["id"],
                "upload_id": o._mapping["upload_id"],
                "order_code": o._mapping.get("order_code") or "",
                "address": addr,
                "worker": o._mapping.get("worker") or "",
                "period_name": o._mapping["period_name"],
                "period_id": o._mapping["period_id"],
                "total": o._mapping.get("total") or 0,
                "order_date": o._mapping.get("order_date"),
                "work_type": get_work_type(o),
                "is_client": bool(o._mapping.get("is_client_payment")),
                "type_label": "ÐšÐ»Ð¸ÐµÐ½Ñ‚" if o._mapping.get("is_client_payment") else "ÐšÐ¾Ð¼Ð¿Ð°Ð½Ð¸Ñ",
            })
        
        # === FIND DUPLICATES (CLUSTER-BASED) ===
        
        exact_duplicates = []
        partial_duplicates = []
        needs_review = []
        seen_review = set()
        
        # Step 1: Group orders by normalized address + work_type
        from collections import defaultdict
        import hashlib
        
        def get_address_key(addr):
            """Create a key for grouping by address"""
            street, number = extract_street_and_number(addr)
            return f"{street}|{number}".lower()
        
        def get_address_hash(addr_key):
            """Create hash for database storage"""
            return hashlib.md5(addr_key.encode()).hexdigest()
        
        # Load existing exclusions - map by (address_hash, work_type) to set of excluded order_ids
        exclusions = await get_duplicate_exclusions()
        if DEBUG_MODE: logger.debug("ðŸ” exclusions from DB: {exclusions}")
        exclusions_map = {}
        for e in exclusions:
            key = (e["address_hash"], e["work_type"])
            excluded_ids = set(e.get("order_ids") or [])
            if DEBUG_MODE: logger.debug("ðŸ” exclusion: key={key}, order_ids={e.get('order_ids')}, excluded_ids={excluded_ids}")
            if key in exclusions_map:
                exclusions_map[key].update(excluded_ids)
            else:
                exclusions_map[key] = excluded_ids
        
        address_groups = defaultdict(list)
        for o in processed_orders:
            # Only consider COMPANY payments for duplicates
            if o["is_client"]:
                continue
            
            key = (get_address_key(o["address"]), o["work_type"])
            address_groups[key].append(o)
        
        # Step 2: Process each group
        for (addr_key, work_type), orders in address_groups.items():
            if len(orders) < 2:
                continue
            
            # Check if ALL orders in group qualify for "repeat visit" exclusion
            # (all have different dates AND all totals < 4000)
            dates = [o.get("order_date") for o in orders]
            totals = [o.get("total") or 0 for o in orders]
            
            # Get unique non-null dates
            unique_dates = set(d for d in dates if d)
            non_null_dates = [d for d in dates if d]
            
            # Debug: log dates for Ð›ÑŽÐ±ÐµÑ€Ñ†Ñ‹
            if "Ð»ÑŽÐ±ÐµÑ€Ñ†" in orders[0]["address"].lower() or "Ð¾ÐºÑ‚ÑÐ±Ñ€ÑŒÑÐºÐ¸Ð¹" in orders[0]["address"].lower():
                if DEBUG_MODE: logger.debug("ðŸ” Ð›ÑŽÐ±ÐµÑ€Ñ†Ñ‹: dates={dates}, unique={unique_dates}, totals={totals}")
            
            # All dates are different (each order on different day)
            all_different_dates = len(unique_dates) == len(non_null_dates) and len(unique_dates) >= 2
            all_small_amounts = all(t < 4000 for t in totals)
            
            if all_different_dates and all_small_amounts:
                # This is likely repeat service visits, not duplicates
                continue
            
            # Check if this cluster is excluded
            address_hash = get_address_hash(addr_key)
            
            # Get current order IDs in this cluster
            current_order_ids = set(o["id"] for o in orders)
            
            # Check if ALL current orders were already marked as "not a duplicate"
            excluded_ids = exclusions_map.get((address_hash, work_type), set())
            if DEBUG_MODE: logger.debug("ðŸ” cluster check: addr_key={addr_key}, hash={address_hash}, work_type={work_type}")
            if DEBUG_MODE: logger.debug("ðŸ” cluster check: current_ids={current_order_ids}, excluded_ids={excluded_ids}")
            if DEBUG_MODE: logger.debug("ðŸ” cluster check: issubset={current_order_ids.issubset(excluded_ids)}")
            if excluded_ids and current_order_ids.issubset(excluded_ids):
                # All orders in this cluster were already checked - skip
                if DEBUG_MODE: logger.debug("ðŸ” cluster SKIPPED")
                continue
            # If there are new orders not in exclusion - show the cluster
            
            # Check if addresses actually match (verify with addresses_match)
            # Use first order's address as reference
            matching_orders = [orders[0]]
            for o in orders[1:]:
                if addresses_match(orders[0]["address"], o["address"]):
                    matching_orders.append(o)
            
            if len(matching_orders) < 2:
                continue
            
            # Determine if exact or partial based on amounts
            amounts = [o["total"] for o in matching_orders]
            all_similar = all(amounts_similar(amounts[0], a) for a in amounts[1:])
            
            # Format orders for output
            formatted_orders = [{
                "id": o["id"],
                "upload_id": o["upload_id"],
                "order_code": o["order_code"],
                "period_name": o["period_name"],
                "worker": o["worker"],
                "total": o["total"],
                "type": o["type_label"],
                "work_type": o["work_type"]
            } for o in matching_orders]
            
            # Sort by period (newest first)
            formatted_orders.sort(key=lambda x: x["period_name"], reverse=True)
            
            cluster = {
                "address": matching_orders[0]["address"],
                "address_hash": address_hash,
                "work_type": work_type,
                "match_type": "exact" if all_similar else "partial",
                "orders": formatted_orders
            }
            
            if all_similar:
                exact_duplicates.append(cluster)
            else:
                partial_duplicates.append(cluster)
        
        # Find needs_review: same order_code, different addresses
        by_order_code = defaultdict(list)
        for o in processed_orders:
            if o["order_code"]:
                by_order_code[o["order_code"]].append(o)
        
        for code, orders_list in by_order_code.items():
            if len(orders_list) < 2:
                continue
            
            # Check if addresses are truly different
            truly_different = False
            for i, o1 in enumerate(orders_list):
                for o2 in orders_list[i+1:]:
                    if not addresses_match(o1["address"], o2["address"]):
                        truly_different = True
                        break
                if truly_different:
                    break
            
            if truly_different:
                ids = tuple(sorted([o["id"] for o in orders_list]))
                if ids not in seen_review:
                    seen_review.add(ids)
                    
                    # Check if this review item is excluded
                    review_hash = 'review_' + code
                    excluded_ids = exclusions_map.get((review_hash, 'review'), set())
                    current_order_ids = set([o["id"] for o in orders_list])
                    
                    if excluded_ids and current_order_ids.issubset(excluded_ids):
                        continue  # Skip excluded review items
                    
                    needs_review.append({
                        "order_code": code,
                        "orders": [{
                            "id": o["id"],
                            "upload_id": o["upload_id"],
                            "address": o["address"],
                            "period_name": o["period_name"],
                            "worker": o["worker"],
                            "total": o["total"],
                            "type": o["type_label"],
                            "work_type": o["work_type"]
                        } for o in orders_list]
                    })
        
        # Sort by period (newest first)
        exact_duplicates.sort(key=lambda x: x["orders"][0]["period_name"], reverse=True)
        partial_duplicates.sort(key=lambda x: x["orders"][0]["period_name"], reverse=True)
        
        return {
            "success": True,
            "exact_duplicates": exact_duplicates,
            "partial_duplicates": partial_duplicates,
            "needs_review": needs_review,
            "stats": {
                "exact_count": len(exact_duplicates),
                "partial_count": len(partial_duplicates),
                "review_count": len(needs_review),
                "total_periods": len(latest_uploads),
                "total_orders": len(processed_orders)
            }
        }
        
    except Exception as e:
        logger.error("âŒ Duplicates API error: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}
@app.get("/duplicates")
async def duplicates_page(request: Request):
    """Duplicate check page (admin only)"""
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/", status_code=302)
    
    return templates.TemplateResponse("duplicates.html", {
        "request": request,
        "user": user,
        "csrf_token": request.state.csrf_token if hasattr(request.state, 'csrf_token') else ''
    })


@app.post("/api/duplicates/exclude")
async def exclude_duplicate(request: Request):
    """Mark a duplicate cluster as 'not a duplicate'"""
    if DEBUG_MODE: logger.debug("ðŸ” exclude_duplicate called")
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        logger.warning("ðŸ” exclude_duplicate: access denied")
        return {"success": False, "error": "Ð”Ð¾ÑÑ‚ÑƒÐ¿ Ð·Ð°Ð¿Ñ€ÐµÑ‰Ñ‘Ð½"}
    
    try:
        data = await request.json()
        if DEBUG_MODE: logger.debug("ðŸ” exclude_duplicate data: {data}")
        address_hash = data.get("address_hash")
        work_type = data.get("work_type")
        address_display = data.get("address_display", "")
        order_ids = data.get("order_ids", [])
        reason = data.get("reason", "")
        
        if not address_hash or not work_type:
            return {"success": False, "error": "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½ Ð°Ð´Ñ€ÐµÑ Ð¸Ð»Ð¸ Ñ‚Ð¸Ð¿ Ñ€Ð°Ð±Ð¾Ñ‚"}
        
        if DEBUG_MODE: logger.debug("ðŸ” Calling add_duplicate_exclusion with order_ids={order_ids}, type={type(order_ids)}")
        exclusion_id = await add_duplicate_exclusion(
            address_hash=address_hash,
            work_type=work_type,
            address_display=address_display,
            order_ids=order_ids,
            excluded_by=user.get("id"),
            excluded_by_name=user.get("display_name", user.get("name", "")),
            reason=reason
        )
        if DEBUG_MODE: logger.debug("ðŸ” add_duplicate_exclusion returned: {exclusion_id}")
        
        return {"success": True, "exclusion_id": exclusion_id}
        
    except Exception as e:
        logger.error("âŒ Exclude duplicate error: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}


@app.delete("/api/duplicates/exclude/{exclusion_id}")
async def restore_duplicate(request: Request, exclusion_id: int):
    """Remove exclusion and show duplicate again"""
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        return {"success": False, "error": "Ð”Ð¾ÑÑ‚ÑƒÐ¿ Ð·Ð°Ð¿Ñ€ÐµÑ‰Ñ‘Ð½"}
    
    try:
        await remove_duplicate_exclusion(exclusion_id)
        return {"success": True}
    except Exception as e:
        logger.error("âŒ Restore duplicate error: {e}")
        return {"success": False, "error": str(e)}


@app.get("/api/duplicates/exclusions")
async def list_exclusions(request: Request):
    """Get list of all excluded duplicates"""
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        return {"success": False, "error": "Ð”Ð¾ÑÑ‚ÑƒÐ¿ Ð·Ð°Ð¿Ñ€ÐµÑ‰Ñ‘Ð½"}
    
    try:
        exclusions = await get_duplicate_exclusions()
        return {"success": True, "exclusions": exclusions}
    except Exception as e:
        logger.error("âŒ List exclusions error: {e}")
        return {"success": False, "error": str(e)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
